import calendar
import datetime
import io
import os, sys, time
import re
import threading
import uuid

os.makedirs("cache", exist_ok=True)
os.makedirs(os.path.join("cache", "logs"), exist_ok=True)

import importlib
import json
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Optional, List, Tuple

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

import urllib3
from dotenv import load_dotenv
from flask import Flask, request, render_template, redirect, url_for, jsonify, \
    send_file
from loguru import logger

# 确保 INFO 级别日志（含 HTTP 计时）在终端输出；级别由 config.xml 的 http_timing.log_level 控制
logger.remove()
logger.add(sys.stderr, level="INFO")
logger.add(
    os.path.join("cache", "logs", "fund_server.log"),
    level="INFO",
    encoding="utf-8",
    rotation="10 MB",
    retention="14 days",
)

import fund
from src.auth import login_required, get_current_user_id, get_current_username, login_user, logout_user
from src.database import Database
from src.module_html import enhance_fund_tab_content
from src.trading_calendar import is_cn_sse_trading_day, iter_cn_sse_trading_days
from src.yaml_config import get_page_refresh_config, load_yaml_config, get_performance_chart_config, get_nav_sync_config, get_server_config

# 从配置加载业绩图表相关常量
_perf_config = get_performance_chart_config()
PERFORMANCE_CHART_INTERVAL_LABELS = _perf_config.get('interval_labels', {
    "ONE_MONTH": "近1月",
    "THREE_MONTH": "近3月",
    "SIX_MONTH": "近6月",
    "ONE_YEAR": "近1年",
    "THREE_YEAR": "近3年",
    "FIVE_YEAR": "近5年",
    "SINCE_ESTABLISHMENT": "成立以来",
})

PERFORMANCE_CHART_INTERVAL_DAYS = _perf_config.get('interval_days', {
    "ONE_MONTH": 31,
    "THREE_MONTH": 93,
    "SIX_MONTH": 186,
    "ONE_YEAR": 365,
    "THREE_YEAR": 365 * 3,
    "FIVE_YEAR": 365 * 5,
})

PERFORMANCE_CHART_INTERVAL_ORDER = _perf_config.get('interval_order', [
    "ONE_MONTH",
    "THREE_MONTH",
    "SIX_MONTH",
    "ONE_YEAR",
    "THREE_YEAR",
    "FIVE_YEAR",
    "SINCE_ESTABLISHMENT",
])

PERFORMANCE_CHART_INTERVALS = set(PERFORMANCE_CHART_INTERVAL_ORDER)
DEFAULT_PERFORMANCE_CHART_INTERVAL = _perf_config.get('default_interval', 'SINCE_ESTABLISHMENT')
DEFAULT_PROFIT_CHART_INTERVAL = _perf_config.get('default_profit_interval', 'THREE_MONTH')

# 从配置加载净值同步相关常量
_nav_config = get_nav_sync_config()
HISTORY_NAV_REQUEST_PAGE_SIZE = _nav_config.get('request_page_size', 300)
NAV_BACKFILL_REQUEST_MONTHS = _nav_config.get('backfill_months', 12)

# 解析当日时间点配置：当日 20:00 前不请求当天净值
_include_today_after_str = _nav_config.get('include_today_after', '20:00')
_nav_include_hour, _nav_include_min = map(int, _include_today_after_str.split(':'))
_NAV_BACKFILL_INCLUDE_TODAY_AFTER = datetime.time(_nav_include_hour, _nav_include_min)


def _nav_backfill_effective_end_date(now: Optional[datetime.datetime] = None) -> datetime.date:
    """历史净值自动补齐所覆盖的最后一日（不含「今天」直至 20:00）。"""
    now = now or datetime.datetime.now()
    today = now.date()
    if now.time() < _NAV_BACKFILL_INCLUDE_TODAY_AFTER:
        return today - datetime.timedelta(days=1)
    return today


def _add_months_keep_day(base_date: datetime.date, months: int) -> datetime.date:
    """为日期加减月份，并尽量保持日数不变。"""
    total_months = (base_date.year * 12 + (base_date.month - 1)) + int(months)
    year = total_months // 12
    month = total_months % 12 + 1
    day = min(base_date.day, calendar.monthrange(year, month)[1])
    return datetime.date(year, month, day)


def _build_backfill_request_segments(
    missing_dates: List[datetime.date],
    max_end_date: datetime.date,
    months: int = NAV_BACKFILL_REQUEST_MONTHS,
) -> List[Tuple[datetime.date, datetime.date]]:
    """按固定月数窗口构造补齐请求片段，窗口长度不受假期缺口影响。"""
    normalized_dates = sorted({
        item for item in missing_dates
        if isinstance(item, datetime.date) and item <= max_end_date
    })
    if not normalized_dates:
        return []

    segments: List[Tuple[datetime.date, datetime.date]] = []
    covered_until: Optional[datetime.date] = None
    for missing_day in normalized_dates:
        if covered_until is not None and missing_day <= covered_until:
            continue
        seg_start = missing_day
        seg_end = min(_add_months_keep_day(seg_start, months) - datetime.timedelta(days=1), max_end_date)
        segments.append((seg_start, seg_end))
        covered_until = seg_end
    return segments


# 加载环境变量
load_dotenv()

urllib3.disable_warnings()
try:
    _ssl_module = getattr(getattr(urllib3, "util", None), "ssl_", None)
    if _ssl_module is not None and hasattr(_ssl_module, "DEFAULT_CIPHERS"):
        _ssl_module.DEFAULT_CIPHERS = ":".join(
            [
                "ECDHE+AESGCM",
                "ECDHE+CHACHA20",
                'ECDHE-RSA-AES128-SHA',
                'ECDHE-RSA-AES256-SHA',
                "RSA+AESGCM",
                'AES128-SHA',
                'AES256-SHA',
            ]
        )
except Exception:
    pass

app = Flask(__name__)
_server_cfg = get_server_config()
app.secret_key = _server_cfg.get('secret_key', 'luobobo')
db = Database()  # 初始化数据库
IMPORT_JOB_STORE = {}
IMPORT_JOB_LOCK = threading.Lock()
IMPORT_DETAIL_LOG_PATH = os.path.join("cache", "logs", "transaction_import.log")
SERVER_LOG_PATH = os.path.join("cache", "logs", "fund_server.log")
LOG_CLEANUP_STATE_PATH = os.path.join("cache", "logs", ".log_cleanup_state")

_LOG_CLEANUP_LOCK = threading.Lock()
_LOG_CLEANUP_THREAD_STARTED = False


def _get_log_cleanup_config():
    """读取日志清理配置，失败时回退默认值。"""
    default_cfg = {
        "enabled": True,
        "retain_days": 14,
        "interval_hours": 6,
    }
    try:
        cfg = load_yaml_config().get("logging_cleanup", {})
        if not isinstance(cfg, dict):
            return default_cfg

        enabled = bool(cfg.get("enabled", default_cfg["enabled"]))

        retain_days = int(cfg.get("retain_days", default_cfg["retain_days"]))
        if retain_days < 1:
            retain_days = default_cfg["retain_days"]

        interval_hours = int(cfg.get("interval_hours", default_cfg["interval_hours"]))
        if interval_hours < 1:
            interval_hours = default_cfg["interval_hours"]

        return {
            "enabled": enabled,
            "retain_days": retain_days,
            "interval_hours": interval_hours,
        }
    except Exception as e:
        logger.warning(f"读取 logging_cleanup 配置失败，使用默认值: {e}")
        return default_cfg


def _parse_log_line_datetime(line):
    """尽量从日志行中解析时间戳。支持 `[YYYY-mm-dd HH:MM:SS]` 与 loguru 默认前缀。"""
    if not line:
        return None

    bracket_match = re.match(r"^\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]", line)
    if bracket_match:
        try:
            return datetime.datetime.strptime(bracket_match.group(1), "%Y-%m-%d %H:%M:%S")
        except Exception:
            pass

    loguru_match = re.match(r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}(?:\.\d+)?)", line)
    if loguru_match:
        ts_text = loguru_match.group(1)
        for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.datetime.strptime(ts_text, fmt)
            except Exception:
                continue

    return None


def _trim_log_file_keep_recent(file_path, retain_days):
    """按日志行时间戳裁剪文件内容，仅保留最近 retain_days 天。"""
    if not os.path.exists(file_path):
        return

    cutoff = datetime.datetime.now() - datetime.timedelta(days=retain_days)

    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()

        kept_lines = []
        dropped_count = 0
        for line in lines:
            dt = _parse_log_line_datetime(line)
            if dt is None or dt >= cutoff:
                kept_lines.append(line)
            else:
                dropped_count += 1

        if dropped_count <= 0:
            return

        temp_path = f"{file_path}.tmp"
        with open(temp_path, "w", encoding="utf-8") as f:
            f.writelines(kept_lines)
        os.replace(temp_path, file_path)

        logger.info(
            f"日志清理完成: file={file_path}, removed_lines={dropped_count}, kept_lines={len(kept_lines)}, retain_days={retain_days}"
        )
    except Exception as e:
        logger.error(f"日志清理失败: file={file_path}, error={e}")


def _run_log_cleanup_once(retain_days):
    """执行一次日志裁剪。"""
    _trim_log_file_keep_recent(SERVER_LOG_PATH, retain_days)
    _trim_log_file_keep_recent(IMPORT_DETAIL_LOG_PATH, retain_days)


def _should_run_log_cleanup(interval_hours):
    """基于持久化状态判断是否需要执行清理，避免重启后频繁触发。"""
    try:
        if not os.path.exists(LOG_CLEANUP_STATE_PATH):
            return True
        last_run_ts = os.path.getmtime(LOG_CLEANUP_STATE_PATH)
        elapsed_seconds = max(0, time.time() - last_run_ts)
        return elapsed_seconds >= max(1, int(interval_hours * 3600))
    except Exception:
        return True


def _mark_log_cleanup_ran():
    """记录最近一次清理时间。"""
    try:
        os.makedirs(os.path.dirname(LOG_CLEANUP_STATE_PATH), exist_ok=True)
        with open(LOG_CLEANUP_STATE_PATH, "w", encoding="utf-8") as state_file:
            state_file.write(datetime.datetime.now().isoformat())
    except Exception as e:
        logger.warning(f"写入日志清理状态失败: {e}")


def _log_cleanup_worker(interval_hours, retain_days):
    """后台定时执行日志清理。"""
    interval_seconds = max(1, int(interval_hours * 3600))
    time.sleep(interval_seconds)
    while True:
        if _should_run_log_cleanup(interval_hours):
            _run_log_cleanup_once(retain_days)
            _mark_log_cleanup_ran()
        time.sleep(interval_seconds)


def _start_log_cleanup_worker_if_needed():
    """启动日志清理线程（仅一次）。"""
    global _LOG_CLEANUP_THREAD_STARTED

    cfg = _get_log_cleanup_config()
    if not cfg.get("enabled", True):
        logger.info("日志清理任务已禁用（logging_cleanup.enabled=false）")
        return

    with _LOG_CLEANUP_LOCK:
        if _LOG_CLEANUP_THREAD_STARTED:
            return

        worker = threading.Thread(
            target=_log_cleanup_worker,
            args=(cfg["interval_hours"], cfg["retain_days"]),
            daemon=True,
            name="log-cleanup-worker",
        )
        worker.start()
        logger.info(
            f"日志清理任务已启动: interval_hours={cfg['interval_hours']}, retain_days={cfg['retain_days']}"
        )
        # 必须最后设置，避免竞态条件
        _LOG_CLEANUP_THREAD_STARTED = True


def _append_import_detail_log(level, message, **fields):
    """写入更适合人工阅读的交易导入本地日志。"""
    try:
        ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        detail_parts = []
        for key, value in fields.items():
            if value is None or value == '':
                continue
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            detail_parts.append(f'{key}={value}')
        detail_text = ' | '.join(detail_parts)
        line = f'[{ts}] [{level}] {message}'
        if detail_text:
            line += f' | {detail_text}'
        with open(IMPORT_DETAIL_LOG_PATH, 'a', encoding='utf-8') as f:
            f.write(line + '\n')
    except Exception as e:
        logger.error(f"写入交易导入本地日志失败: {e}")


def _set_import_job_state(job_id, **kwargs):
    with IMPORT_JOB_LOCK:
        state = IMPORT_JOB_STORE.get(job_id, {}).copy()
        state.update(kwargs)
        state['updated_at'] = time.time()
        IMPORT_JOB_STORE[job_id] = state
        return state


def _get_import_job_state(job_id):
    with IMPORT_JOB_LOCK:
        state = IMPORT_JOB_STORE.get(job_id)
        return dict(state) if state else None


def _cleanup_finished_import_jobs(max_age_seconds=1800):
    now_ts = time.time()
    with IMPORT_JOB_LOCK:
        expired_ids = [
            job_id for job_id, state in IMPORT_JOB_STORE.items()
            if state.get('done') and (now_ts - float(state.get('updated_at', now_ts))) > max_age_seconds
        ]
        for job_id in expired_ids:
            IMPORT_JOB_STORE.pop(job_id, None)


# ==================== Authentication Routes ====================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """登录页面和处理"""
    if request.method == 'GET':
        # 检查是否有记住我的cookie
        remember_token = request.cookies.get('remember_token')
        if remember_token:
            # 尝试从token中解析用户信息并自动登录
            try:
                import hashlib
                # token格式: username:hashed_password
                parts = remember_token.split(':')
                if len(parts) == 2:
                    username, token_hash = parts
                    user = db.get_user_by_username(username)
                    if user:
                        # 验证token是否匹配
                        expected_hash = hashlib.sha256(f"{username}:{user['password_hash']}".encode()).hexdigest()
                        if token_hash == expected_hash:
                            login_user(user['id'], username)
                            return redirect(url_for('get_fund'))
            except Exception as e:
                logger.error(f"Auto-login failed: {e}")

        return render_template('login.html')

    # POST request - handle login
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    remember_me = request.form.get('remember_me') == '1'

    if not username or not password:
        return render_template('login.html', error='请输入用户名和密码')

    success, user_id = db.verify_password(username, password)
    if success:
        login_user(user_id, username)
        response = redirect(url_for('get_fund'))

        # 如果勾选了记住我，设置cookie（7天有效）
        if remember_me:
            import hashlib
            user = db.get_user_by_username(username)
            if not user:
                return response
            token_hash = hashlib.sha256(f"{username}:{user['password_hash']}".encode()).hexdigest()
            remember_token = f"{username}:{token_hash}"
            response.set_cookie('remember_token', remember_token, max_age=7 * 24 * 60 * 60, httponly=True,
                                samesite='Lax')

        return response
    else:
        return render_template('login.html', error='用户名或密码错误')


@app.route('/register', methods=['GET', 'POST'])
def register():
    """注册页面和处理"""
    if request.method == 'GET':
        return render_template('register.html')

    # POST request - handle registration
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    confirm_password = request.form.get('confirm_password', '')

    # 验证输入
    if not username or not password:
        return render_template('register.html', error='请输入用户名和密码')

    if len(username) < 3 or len(username) > 20:
        return render_template('register.html', error='用户名长度应为3-20个字符')

    if len(password) < 6:
        return render_template('register.html', error='密码长度至少为6个字符')

    if password != confirm_password:
        return render_template('register.html', error='两次输入的密码不一致')

    # 创建用户
    success, message, user_id = db.create_user(username, password)
    if success:
        # 注册成功，自动登录
        login_user(user_id, username)
        return redirect(url_for('get_fund'))
    else:
        return render_template('register.html', error=message)


@app.route('/logout')
def logout():
    """登出"""
    logout_user()
    response = redirect(url_for('login'))
    # 清除记住我的cookie
    response.set_cookie('remember_token', '', max_age=0)
    return response


@app.route('/fund/sector', methods=['GET'])
@login_required
def get_sector_funds():
    """获取指定板块的基金列表"""
    bk_id = request.args.get("bk_id")
    importlib.reload(fund)
    my_fund = fund.LanFund(db=db)
    return my_fund.select_fund_html(bk_id=bk_id)


# API endpoints for fund operations
@app.route('/api/fund/add', methods=['POST'])
@login_required
def api_fund_add():
    """添加基金"""
    start = time.perf_counter()
    try:
        data = request.json
        codes = data.get('codes', '')
        if not codes:
            return {'success': False, 'message': '请提供基金代码'}
        user_id = get_current_user_id()
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=user_id, db=db)
        my_fund.add_code(codes)
        result = {'success': True, 'message': f'已添加基金: {codes}'}
        return result
    except Exception as e:
        logger.error(f"添加基金失败: {e}")
        result = {'success': False, 'message': f'添加失败: {str(e)}'}
        return result
    finally:
        elapsed = (time.perf_counter() - start) * 1000
        logger.info(f"[API] /api/fund/add elapsed_ms={elapsed:.1f}")


@app.route('/api/fund/backfill-establishment-dates', methods=['POST'])
@login_required
def api_backfill_establishment_dates():
    """批量补齐当前用户基金成立日期（仅处理 establishment_date 缺失的数据）。"""
    start = time.perf_counter()
    try:
        user_id = get_current_user_id()
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=user_id, db=db)

        fund_map = my_fund.CACHE_MAP if isinstance(my_fund.CACHE_MAP, dict) else {}
        if not fund_map:
            return {
                'success': True,
                'message': '暂无基金可回填',
                'total': 0,
                'missing': 0,
                'updated': 0,
                'failed': 0,
                'failed_codes': [],
            }

        total_count = len(fund_map)
        missing_codes = []
        updated_count = 0
        failed_codes = []

        for fund_code, fund_data in fund_map.items():
            existing_text = my_fund._normalize_establishment_date_text(
                (fund_data or {}).get('establishment_date') if isinstance(fund_data, dict) else ''
            )
            if existing_text:
                continue

            missing_codes.append(fund_code)
            established = my_fund._ensure_fund_establishment_date(fund_code)
            if isinstance(established, datetime.date):
                updated_count += 1
            else:
                failed_codes.append(fund_code)

        missing_count = len(missing_codes)
        failed_count = len(failed_codes)
        result = {
            'success': True,
            'message': f'成立日回填完成：缺失{missing_count}，补齐{updated_count}，失败{failed_count}',
            'total': total_count,
            'missing': missing_count,
            'updated': updated_count,
            'failed': failed_count,
            'failed_codes': failed_codes,
        }
        return result
    except Exception as e:
        logger.error(f"批量回填成立日期失败: {e}")
        return {
            'success': False,
            'message': f'回填失败: {str(e)}',
            'total': 0,
            'missing': 0,
            'updated': 0,
            'failed': 0,
            'failed_codes': [],
        }
    finally:
        elapsed = (time.perf_counter() - start) * 1000
        logger.info(f"[API] /api/fund/backfill-establishment-dates elapsed_ms={elapsed:.1f}")


@app.route('/api/fund/delete', methods=['POST'])
@login_required
def api_fund_delete():
    """删除基金"""
    start = time.perf_counter()
    try:
        data = request.json
        codes = data.get('codes', '')
        if not codes:
            return {'success': False, 'message': '请提供基金代码'}
        user_id = get_current_user_id()
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=user_id, db=db)
        my_fund.delete_code(codes)
        result = {'success': True, 'message': f'已删除基金: {codes}'}
        return result
    except Exception as e:
        logger.error(f"删除基金失败: {e}")
        result = {'success': False, 'message': f'删除失败: {str(e)}'}
        return result
    finally:
        elapsed = (time.perf_counter() - start) * 1000
        logger.info(f"[API] /api/fund/delete elapsed_ms={elapsed:.1f}")


@app.route('/api/fund/hold', methods=['POST'])
@login_required
def api_fund_set_hold():
    """设置/取消持有标记"""
    start = time.perf_counter()
    try:
        data = request.json
        codes = data.get('codes', '')
        hold = data.get('hold', True)
        if not codes:
            return {'success': False, 'message': '请提供基金代码'}
        user_id = get_current_user_id()
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=user_id, db=db)
        code_list = [c.strip() for c in codes.split(',')]
        for code in code_list:
            if code in my_fund.CACHE_MAP:
                my_fund.CACHE_MAP[code]['is_hold'] = hold
        my_fund.save_cache()
        action = '标记持有' if hold else '取消持有'
        result = {'success': True, 'message': f'已{action}: {codes}'}
        return result
    except Exception as e:
        logger.error(f"设置持有标记失败: {e}")
        result = {'success': False, 'message': f'操作失败: {str(e)}'}
        return result
    finally:
        elapsed = (time.perf_counter() - start) * 1000
        logger.info(f"[API] /api/fund/hold elapsed_ms={elapsed:.1f}")


@app.route('/api/fund/sector', methods=['POST'])
@login_required
def api_fund_set_sector():
    """设置板块标记"""
    start = time.perf_counter()
    try:
        data = request.json
        codes = data.get('codes', '')
        sectors = data.get('sectors', [])
        if not codes:
            return {'success': False, 'message': '请提供基金代码'}
        if not sectors:
            return {'success': False, 'message': '请选择板块'}
        user_id = get_current_user_id()
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=user_id, db=db)
        code_list = [c.strip() for c in codes.split(',')]
        # 使用Web专用方法
        my_fund.mark_fund_sector_web(code_list, sectors)
        sectors_str = ", ".join(sectors)
        result = {'success': True, 'message': f'已标注板块: {codes} -> {sectors_str}'}
        return result
    except Exception as e:
        logger.error(f"标注板块失败: {e}")
        result = {'success': False, 'message': f'操作失败: {str(e)}'}
        return result
    finally:
        elapsed = (time.perf_counter() - start) * 1000
        logger.info(f"[API] /api/fund/sector elapsed_ms={elapsed:.1f}")


@app.route('/api/fund/sector/remove', methods=['POST'])
@login_required
def api_fund_remove_sector():
    """删除板块标记"""
    start = time.perf_counter()
    try:
        data = request.json
        codes = data.get('codes', '')
        if not codes:
            return {'success': False, 'message': '请提供基金代码'}
        user_id = get_current_user_id()
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=user_id, db=db)
        code_list = [c.strip() for c in codes.split(',')]
        # 使用Web专用方法
        my_fund.unmark_fund_sector_web(code_list)
        result = {'success': True, 'message': f'已删除板块标记: {codes}'}
        return result
    except Exception as e:
        logger.error(f"删除板块标记失败: {e}")
        result = {'success': False, 'message': f'操作失败: {str(e)}'}
        return result
    finally:
        elapsed = (time.perf_counter() - start) * 1000
        logger.info(f"[API] /api/fund/sector/remove elapsed_ms={elapsed:.1f}")


# ==================== fund_map File Upload/Download ====================

@app.route('/api/fund/upload', methods=['POST'])
@login_required
def api_fund_upload():
    """上传fund_map.json文件"""
    start = time.perf_counter()
    try:
        if 'file' not in request.files:
            return {'success': False, 'message': '未找到上传文件'}

        file = request.files['file']
        filename = str(file.filename or '').strip()
        if filename == '':
            return {'success': False, 'message': '未选择文件'}

        if not filename.endswith('.json'):
            return {'success': False, 'message': '只支持JSON文件'}

        # 读取并解析JSON
        content = file.read().decode('gbk')  # 使用GBK编码
        fund_map = json.loads(content)

        # 验证数据格式
        if not isinstance(fund_map, dict):
            return {'success': False, 'message': '文件格式错误：应为JSON对象'}

        for code, fund_data in fund_map.items():
            if not isinstance(fund_data, dict):
                return {'success': False, 'message': f'基金{code}数据格式错误'}
            if 'fund_key' not in fund_data or 'fund_name' not in fund_data:
                return {'success': False, 'message': f'基金{code}缺少必要字段'}

        # 保存到数据库
        user_id = get_current_user_id()
        success = db.save_user_funds(user_id, fund_map)

        if success:
            result = {'success': True, 'message': f'成功导入{len(fund_map)}个基金'}
        else:
            result = {'success': False, 'message': '保存失败'}
        return result

    except json.JSONDecodeError:
        result = {'success': False, 'message': 'JSON格式错误'}
        return result
    except Exception as e:
        logger.error(f"上传文件失败: {e}")
        result = {'success': False, 'message': f'上传失败: {str(e)}'}
        return result
    finally:
        elapsed = (time.perf_counter() - start) * 1000
        logger.info(f"[API] /api/fund/upload elapsed_ms={elapsed:.1f}")


@app.route('/api/fund/download', methods=['GET'])
@login_required
def api_fund_download():
    """下载fund_map.json文件"""
    try:
        user_id = get_current_user_id()
        fund_map = db.get_user_funds(user_id)

        # 生成JSON文件
        import tempfile
        with tempfile.NamedTemporaryFile(mode='w', encoding='gbk', suffix='.json', delete=False) as f:
            json.dump(fund_map, f, ensure_ascii=False, indent=4)
            temp_path = f.name

        return send_file(
            temp_path,
            as_attachment=True,
            download_name='fund_map.json',
            mimetype='application/json'
        )

    except Exception as e:
        logger.error(f"下载文件失败: {e}")
        return {'success': False, 'message': f'下载失败: {str(e)}'}


@app.route('/api/fund/transactions/download-all', methods=['GET'])
@login_required
def api_fund_transactions_download_all():
    """下载当前用户的全部交易记录备份。"""
    try:
        user_id = get_current_user_id()
        user_funds = db.get_user_funds(user_id)
        transactions = db.get_all_fund_transactions(user_id)

        backup_payload = {
            'exported_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'user_id': user_id,
            'fund_count': len(user_funds),
            'transaction_count': len(transactions),
            'funds': user_funds,
            'transactions': transactions,
        }

        import tempfile
        with tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', suffix='.json', delete=False) as f:
            json.dump(backup_payload, f, ensure_ascii=False, indent=2)
            temp_path = f.name

        download_name = f'fund_transactions_backup_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        return send_file(
            temp_path,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/json'
        )
    except Exception as e:
        logger.error(f"下载全部交易记录备份失败: {e}")
        return {'success': False, 'message': f'下载失败: {str(e)}'}


# ==================== Shares Management ====================


def _normalize_nav_date(nav_date_text):
    """规范化净值日期到 YYYY-MM-DD。"""
    if not nav_date_text:
        return None
    nav_date_text = str(nav_date_text).strip()
    try:
        if len(nav_date_text) == 5:  # MM-DD
            current_year = datetime.date.today().year
            nav_date_text = f"{current_year}-{nav_date_text}"
        return datetime.date.fromisoformat(nav_date_text).isoformat()
    except Exception:
        return None


def _extract_net_value_and_date(net_value_text):
    """从形如 1.2345(2026-03-15) 的字符串提取净值和日期。"""
    text = str(net_value_text or '').strip()
    if not text:
        return None, None

    nav_value = None
    nav_date = None
    try:
        nav_value = float(text.split('(')[0])
    except Exception:
        nav_value = None

    if '(' in text and ')' in text:
        try:
            date_text = text.split('(')[1].split(')')[0]
            nav_date = _normalize_nav_date(date_text)
        except Exception:
            nav_date = None

    return nav_value, nav_date


def _get_latest_fund_quote(user_id, fund_code):
    """获取基金最新净值与净值日期。"""
    user_funds = db.get_user_funds(user_id)
    if fund_code not in user_funds:
        return None, None, None

    # 最新净值不是从业绩曲线接口取，而是复用 fund.search_code(True) 的结果。
    # 该结果内部会请求 DATA_SOURCE_URLS['fund123_matiaria_tpl']，并从响应中解析
    # netValue / netValueDate，这也是页面基金表“最新净值”的来源。
    importlib.reload(fund)
    my_fund = fund.LanFund(user_id=user_id, db=db)
    rows = my_fund.search_code(True) or []
    for row in rows:
        if row[0] == fund_code:
            net_value, nav_date = _extract_net_value_and_date(row[3])
            return net_value, nav_date, user_funds[fund_code]
    return None, None, user_funds[fund_code]


def _get_buy_effective_date(now_time=None):
    """买入生效净值日：15:00前为当日，15:00后为次日（交易日由净值更新自然对齐）。"""
    now_time = now_time or datetime.datetime.now()
    effective_date = now_time.date()
    if now_time.hour >= 15:
        effective_date = effective_date + datetime.timedelta(days=1)
    return effective_date.isoformat()


def _settle_pending_buys(user_id):
    """结算到期的待确认买入单。"""
    pending_orders = db.get_pending_buys(user_id)
    if not pending_orders:
        return 0

    settled_count = 0
    quote_cache = {}

    for order in pending_orders:
        fund_code = str(order.get('fund_code', '')).strip()
        if not fund_code:
            continue

        effective_date_text = str(order.get('effective_date', '')).strip()
        try:
            effective_date = datetime.date.fromisoformat(effective_date_text)
        except Exception:
            continue

        if fund_code not in quote_cache:
            quote_cache[fund_code] = _get_latest_fund_quote(user_id, fund_code)
        latest_net_value, latest_nav_date_text, _fund_data = quote_cache[fund_code]

        if not latest_net_value or latest_net_value <= 0 or not latest_nav_date_text:
            continue

        try:
            latest_nav_date = datetime.date.fromisoformat(latest_nav_date_text)
        except Exception:
            continue

        if latest_nav_date < effective_date:
            continue

        amount = float(order.get('amount', 0) or 0)
        shares = float((Decimal(str(amount)) / Decimal(str(latest_net_value))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        if amount <= 0 or shares <= 0:
            continue

        current_shares = db.update_fund_shares_delta(user_id, fund_code, shares)
        if current_shares is None:
            continue

        tx_time = f"{effective_date.isoformat()} 15:00:00"
        tx_id = db.add_fund_transaction(
            user_id=user_id,
            fund_code=fund_code,
            tx_type='buy',
            amount=amount,
            shares=shares,
            net_value=latest_net_value,
            tx_time=tx_time,
            fee=0,
        )

        if tx_id is None:
            db.update_fund_shares_delta(user_id, fund_code, -shares)
            continue

        marked = db.mark_pending_buy_settled(
            pending_id=order['id'],
            settled_tx_id=tx_id,
            settled_net_value=latest_net_value,
            settled_shares=shares,
        )
        if marked:
            settled_count += 1

    return settled_count


def _find_net_value_by_date_from_trend(user_id, fund_code, target_date):
    """从基金业绩趋势点中查找指定日期净值（若趋势数据包含净值）。"""
    user_funds = db.get_user_funds(user_id)
    fund_data = user_funds.get(fund_code)
    if not fund_data:
        return None

    importlib.reload(fund)
    my_fund = fund.LanFund(user_id=user_id, db=db)
    chart_data = my_fund.get_fund_performance_chart_data(fund_code, fund_data, "SINCE_ESTABLISHMENT")

    labels = chart_data.get('labels', []) or []
    net_values = chart_data.get('net_values', []) or []
    if not labels or not net_values:
        return None

    short_date = target_date[5:] if isinstance(target_date, str) and len(target_date) == 10 else target_date
    for index, label in enumerate(labels):
        if label != target_date and label != short_date:
            continue
        if index >= len(net_values):
            continue
        value = net_values[index]
        try:
            nav = float(value)
            if nav > 0:
                return round(nav, 4)
        except (TypeError, ValueError):
            continue

    return None


def _find_net_value_by_date_from_history_api(user_id, fund_code, target_date):
    """通过历史净值接口按指定日期精确查询净值。"""
    user_funds = db.get_user_funds(user_id)
    fund_data = user_funds.get(fund_code)
    if not fund_data:
        return None

    fund_key = fund_data.get('fund_key')
    if not fund_key:
        return None

    try:
        normalized_date = datetime.date.fromisoformat(str(target_date)).strftime("%Y%m%d")
    except Exception:
        return None

    importlib.reload(fund)
    my_fund = fund.LanFund(user_id=user_id, db=db)

    api_url = fund.DATA_SOURCE_URLS.get('fund123_history_net_value_api')
    if not api_url:
        return None

    headers = {
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        "Content-Type": "application/json",
        "Origin": fund.DATA_SOURCE_URLS['fund123_origin'],
        "Referer": fund.DATA_SOURCE_URLS['fund123_fund_page'],
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
        "X-API-Key": "foobar",
        "accept": "json"
    }

    payload = {
        "productId": fund_key,
        "startDate": normalized_date,
        "endDate": normalized_date,
        "pageNum": 1,
        "pageSize": 10,
    }

    try:
        response = my_fund.session.post(
            api_url,
            params={"_csrf": my_fund._csrf},
            json=payload,
            headers=headers,
            timeout=10,
            verify=False,
        )
        response_json = response.json()
    except Exception as e:
        logger.warning(f"历史净值接口请求失败【{fund_code} {target_date}】: {e}")
        return None

    if not response_json.get("success"):
        return None

    value_list = response_json.get("list", []) or []
    if not value_list:
        return None

    for item in value_list:
        item_date = str(item.get("netValueDate", "")).strip()
        if item_date and item_date != str(target_date):
            continue
        try:
            net_value = float(item.get("netValue"))
            if net_value > 0:
                return round(net_value, 4)
        except (TypeError, ValueError):
            continue
    return None


def _get_fund_establishment_date(user_id, fund_code):
    """从 fund123_matiaria 接口读取基金成立日期。"""
    try:
        api_tpl = fund.DATA_SOURCE_URLS.get('fund123_matiaria_tpl')
        if not api_tpl:
            return None

        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=user_id, db=db)
        api_url = api_tpl.format(fund=fund_code)
        response = my_fund.session.get(api_url, timeout=10, verify=False)

        date_text = None
        try:
            payload = response.json()
            date_text = (
                payload.get('titleInfo', {})
                .get('establishmentDate')
            )
        except Exception:
            pass

        if not date_text:
            match = re.search(r'"establishmentDate"\s*:\s*"(\d{4}-\d{2}-\d{2}|\d{8})"', response.text)
            if match:
                date_text = match.group(1)

        if not date_text:
            return None

        date_text = str(date_text).strip()
        if re.fullmatch(r'\d{8}', date_text):
            return datetime.datetime.strptime(date_text, '%Y%m%d').date()
        return datetime.date.fromisoformat(date_text)
    except Exception:
        return None


def _resolve_curve_label_dates(labels: List[str]) -> List[Optional[datetime.date]]:
    """将业绩曲线标签解析为日期；兼容 MM-DD（按时间序列回推年份）。"""
    resolved: List[Optional[datetime.date]] = []
    if not labels:
        return resolved

    has_full_date = any(re.fullmatch(r'\d{4}-\d{2}-\d{2}', str(item or '').strip()) for item in labels)
    if has_full_date:
        for label in labels:
            try:
                resolved.append(datetime.date.fromisoformat(str(label).strip()))
            except Exception:
                resolved.append(None)
        return resolved

    # 仅有 MM-DD 时：从尾部向前回推年份，避免全部落到当年
    current_year = datetime.date.today().year
    reverse_dates: List[Optional[datetime.date]] = []
    prev_date: Optional[datetime.date] = None
    for label in reversed(labels):
        text = str(label or '').strip()
        if not re.fullmatch(r'\d{2}-\d{2}', text):
            reverse_dates.append(None)
            continue
        month = int(text[:2])
        day = int(text[3:5])
        candidate_year = current_year if prev_date is None else prev_date.year
        candidate = None
        for year in range(candidate_year, candidate_year - 6, -1):
            try:
                temp = datetime.date(year, month, day)
            except Exception:
                continue
            if prev_date is None or temp <= prev_date:
                candidate = temp
                break
        reverse_dates.append(candidate)
        if candidate is not None:
            prev_date = candidate

    reverse_dates.reverse()
    return reverse_dates


def _parse_iso_date(date_text):
    try:
        text = str(date_text or '').strip()
        if not text:
            return None
        return datetime.date.fromisoformat(text)
    except Exception:
        return None


def _get_local_fund_establishment_date(user_id, fund_code, user_funds=None):
    """优先从 user_funds 读取成立日，其次用本地净值最早日期推导。"""
    if user_funds is None:
        user_funds = db.get_user_funds(user_id)

    fund_data = user_funds.get(fund_code, {}) if isinstance(user_funds, dict) else {}
    established = _parse_iso_date(fund_data.get('establishment_date'))
    if established:
        return established

    # 成立日缺失时尝试远端补齐，避免误判可用区间（例如成立超 1 年但“近1年”按钮缺失）。
    remote_established = _get_fund_establishment_date(user_id, fund_code)
    if isinstance(remote_established, datetime.date):
        try:
            db.update_fund_establishment_date(user_id, fund_code, remote_established.isoformat())
        except Exception:
            pass
        return remote_established

    nav_map = db.get_fund_nav_history_range(fund_code) or {}
    first_nav_date = None
    for day in sorted(nav_map.keys()):
        day_obj = _parse_iso_date(day)
        if day_obj is not None:
            first_nav_date = day_obj
            break

    # 注意：此处刻意不写入 DB，仅作运行时回退值使用。
    # 若将本地净值最早日误写入 establishment_date，后续调用会把它当成真实成立日，
    # 导致跳过远端 API 获取，区间按钮和净值补齐均会出错。
    return first_nav_date


def _get_available_performance_intervals(establishment_date, end_date):
    available_keys = []
    if not isinstance(end_date, datetime.date):
        end_date = datetime.date.today()

    if isinstance(establishment_date, datetime.date):
        age_days = max((end_date - establishment_date).days, 0)
        for interval_key in PERFORMANCE_CHART_INTERVAL_ORDER:
            if interval_key == 'SINCE_ESTABLISHMENT':
                available_keys.append(interval_key)
                continue
            threshold = PERFORMANCE_CHART_INTERVAL_DAYS.get(interval_key)
            if threshold is not None and age_days >= threshold:
                available_keys.append(interval_key)
        if not available_keys:
            available_keys = ['SINCE_ESTABLISHMENT']
    else:
        available_keys = [
            key for key in PERFORMANCE_CHART_INTERVAL_ORDER
            if key != 'SINCE_ESTABLISHMENT'
        ]

    # 无论是否拿到成立日，统一提供“成立以来”：
    # - 有成立日：从成立日开始
    # - 无成立日：回退到本地最早净值日（保证默认区间可用且数据尽量完整）
    if 'SINCE_ESTABLISHMENT' not in available_keys:
        available_keys.append('SINCE_ESTABLISHMENT')

    return [[key, PERFORMANCE_CHART_INTERVAL_LABELS.get(key, key)] for key in available_keys]


def _get_interval_start_date(interval_key, end_date, establishment_date=None, fallback_start=None):
    if not isinstance(end_date, datetime.date):
        return fallback_start

    if interval_key == 'SINCE_ESTABLISHMENT':
        return establishment_date or fallback_start

    interval_days = PERFORMANCE_CHART_INTERVAL_DAYS.get(interval_key)
    if interval_days is None:
        return fallback_start

    start_date = end_date - datetime.timedelta(days=max(interval_days - 1, 0))
    if isinstance(establishment_date, datetime.date) and start_date < establishment_date:
        start_date = establishment_date
    return start_date


def _ensure_nav_history_from_establishment(user_id, fund_code, establishment_date):
    """确保本地净值覆盖「成立日 -> 有效截止日」区间（增量补齐）；当日 20:00 前不补今天。"""
    if not isinstance(establishment_date, datetime.date):
        return db.get_fund_nav_history_range(fund_code) or {}

    today = datetime.date.today()
    if establishment_date > today:
        return db.get_fund_nav_history_range(fund_code) or {}

    backfill_end = _nav_backfill_effective_end_date()
    if establishment_date > backfill_end:
        return db.get_fund_nav_history_range(
            fund_code,
            establishment_date.isoformat(),
            today.isoformat(),
        ) or {}

    local_nav_map = db.get_fund_nav_history_range(
        fund_code,
        establishment_date.isoformat(),
        today.isoformat(),
    ) or {}

    # 获取本地最新净值的日期，作为增量补齐的起点
    sorted_dates = sorted(local_nav_map.keys())
    local_max_date = sorted_dates[-1] if sorted_dates else None

    # 判断是否需要完整补齐
    # 规则：如果本地数据覆盖范围 < 成立日到 backfill_end 的 80%，认为是首次或数据严重缺失，需要全量补齐
    establishment_to_end_days = len(iter_cn_sse_trading_days(establishment_date, backfill_end))
    local_data_days = len(sorted_dates)

    need_full_backfill = not local_max_date or local_data_days < establishment_to_end_days * 0.8

    logger.info(
        f"净值补齐检查【{fund_code}】: establishment_date={establishment_date}, backfill_end={backfill_end}, "
        f"establishment_to_end_days={establishment_to_end_days}, local_data_days={local_data_days}, "
        f"local_max_date={local_max_date}, threshold_80%={int(establishment_to_end_days * 0.8)}, "
        f"need_full_backfill={need_full_backfill}"
    )

    # 确定检查范围并找出缺失交易日
    # 首次请求 或 本地数据不完整：从成立日开始检查完整缺失
    if need_full_backfill:
        # 首次请求：检查从成立日到 backfill_end 是否有缺失
        check_start = establishment_date
        all_days = iter_cn_sse_trading_days(establishment_date, backfill_end)
        missing_days = [d for d in all_days if d.isoformat() not in local_nav_map]
    else:
        # 数据已完整：增量检查本地最新日期的下一个自然日之后是否有缺失交易日
        local_max_date_obj = datetime.date.fromisoformat(local_max_date)
        check_start = local_max_date_obj + datetime.timedelta(days=1)
        all_days = iter_cn_sse_trading_days(check_start, backfill_end)
        missing_days = [d for d in all_days if d.isoformat() not in local_nav_map]

    if not missing_days:
        # 没有缺失交易日，无需请求
        logger.debug(f"净值补齐跳过【{fund_code}】: 本地已有 {len(sorted_dates)} 天数据，从 {local_max_date or establishment_date} 起无缺失交易日")
        return local_nav_map

    # 构建请求分段
    request_segments = _build_backfill_request_segments(missing_days, backfill_end)
    segment_count = len(request_segments)

    if segment_count == 0:
        return local_nav_map

    # 调试日志：显示为什么需要补齐
    logger.info(
        f"净值补齐【{fund_code}】: start={establishment_date}, backfill_end={backfill_end}, today={today}, "
        f"local_max_date={local_max_date}, missing_days={len(missing_days)}, segments={segment_count}"
    )

    started_at = time.perf_counter()
    wrote_count = 0
    fetched_count = 0
    for seg_start, seg_end in request_segments:
        if seg_start > seg_end:
            continue
        remote_nav_map = _fetch_history_nav_map_by_date_range(
            user_id,
            fund_code,
            seg_start.isoformat(),
            seg_end.isoformat(),
        )
        if not remote_nav_map:
            continue
        fetched_count += len(remote_nav_map)
        for nav_date, nav_value in remote_nav_map.items():
            if db.upsert_fund_nav_history(fund_code, nav_date, nav_value, source='history_api_establishment_sync'):
                wrote_count += 1

    elapsed_ms = int((time.perf_counter() - started_at) * 1000)
    logger.info(
        f"净值补齐【{fund_code}】: start={establishment_date}, backfill_end={backfill_end}, today={today}, segments={segment_count}, fetched={fetched_count}, wrote={wrote_count}, elapsed_ms={elapsed_ms}"
    )

    return db.get_fund_nav_history_range(
        fund_code,
        establishment_date.isoformat(),
        today.isoformat(),
    ) or local_nav_map


def _build_local_performance_chart_data(user_id, fund_code, fund_name, date_interval=DEFAULT_PERFORMANCE_CHART_INTERVAL, user_funds=None):
    """基于本地 fund_nav_history 构建业绩曲线，不依赖远端业绩曲线接口。"""
    if user_funds is None:
        user_funds = db.get_user_funds(user_id)

    # 先读取成立日（缺失时补齐），再确保本地净值覆盖“成立日 -> 今日”后再计算曲线。
    establishment_date = _get_local_fund_establishment_date(user_id, fund_code, user_funds=user_funds)
    if isinstance(establishment_date, datetime.date):
        nav_map = _ensure_nav_history_from_establishment(user_id, fund_code, establishment_date)
    else:
        nav_map = db.get_fund_nav_history_range(fund_code) or {}

    # 统一从本地净值历史取数，作为业绩与收益曲线的共同底座。
    nav_points = []
    for day in sorted(nav_map.keys()):
        day_obj = _parse_iso_date(day)
        nav_value = _safe_float(nav_map.get(day), None)
        if day_obj is None or nav_value is None or nav_value <= 0:
            continue
        nav_points.append((day_obj, round(nav_value, 4)))

    end_date = nav_points[-1][0] if nav_points else datetime.date.today()
    available_intervals = _get_available_performance_intervals(establishment_date, end_date)
    available_keys = [item[0] for item in available_intervals]

    selected_interval = str(date_interval or DEFAULT_PERFORMANCE_CHART_INTERVAL).strip().upper()
    if selected_interval not in available_keys:
        if DEFAULT_PERFORMANCE_CHART_INTERVAL in available_keys:
            selected_interval = DEFAULT_PERFORMANCE_CHART_INTERVAL
        elif 'SINCE_ESTABLISHMENT' in available_keys:
            selected_interval = 'SINCE_ESTABLISHMENT'
        elif available_keys:
            selected_interval = available_keys[0]
        else:
            selected_interval = DEFAULT_PERFORMANCE_CHART_INTERVAL

    if nav_points:
        fallback_start = nav_points[0][0]
        start_date = _get_interval_start_date(
            selected_interval,
            end_date,
            establishment_date=establishment_date,
            fallback_start=fallback_start,
        )
        filtered_points = [item for item in nav_points if (start_date is None or item[0] >= start_date)]
        if not filtered_points:
            filtered_points = nav_points
    else:
        filtered_points = []

    labels = [item[0].isoformat() for item in filtered_points]
    net_values = [item[1] for item in filtered_points]

    growth = []
    base_nav = net_values[0] if net_values else None
    for nav in net_values:
        if base_nav is None or base_nav <= 0:
            growth.append(None)
        else:
            growth.append(round((nav / base_nav - 1.0) * 100.0, 2))

    latest_net_value = net_values[-1] if net_values else None
    latest_net_value_date = labels[-1] if labels else None

    return {
        'labels': labels,
        'growth': growth,
        'net_values': net_values,
        'benchmark_label': '沪深300',
        'benchmark_growth': [],
        'latest_net_value': latest_net_value,
        'latest_net_value_date': latest_net_value_date,
        'from_cache': True,
        'date_interval': selected_interval,
        'interval_label': PERFORMANCE_CHART_INTERVAL_LABELS.get(selected_interval, selected_interval),
        'available_intervals': available_intervals,
        'establishment_date': establishment_date.isoformat() if isinstance(establishment_date, datetime.date) else None,
        'growth_source': 'local_nav',
    }


def _fetch_history_nav_map_by_date_range(user_id, fund_code, start_date, end_date):
    """通过历史净值接口批量获取区间净值，返回 {YYYY-MM-DD: nav}。"""
    result = {}
    try:
        start_obj = datetime.date.fromisoformat(str(start_date))
        end_obj = datetime.date.fromisoformat(str(end_date))
    except Exception:
        return result

    if start_obj > end_obj:
        return result

    user_funds = db.get_user_funds(user_id)
    fund_data = user_funds.get(fund_code)
    if not fund_data:
        return result

    fund_key = fund_data.get('fund_key')
    if not fund_key:
        return result

    api_url = fund.DATA_SOURCE_URLS.get('fund123_history_net_value_api')
    if not api_url:
        return result

    my_fund = fund.LanFund(user_id=user_id, db=db)

    headers = {
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        "Content-Type": "application/json",
        "Origin": fund.DATA_SOURCE_URLS['fund123_origin'],
        "Referer": fund.DATA_SOURCE_URLS['fund123_fund_page'],
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
        "X-API-Key": "foobar",
        "accept": "json"
    }

    page_size = HISTORY_NAV_REQUEST_PAGE_SIZE
    max_pages = 200

    request_started_at = time.perf_counter()
    segment_total = 0
    page_total = 0

    # 接口限制单次 start/end 不超过 600 日
    segment_start = start_obj
    while segment_start <= end_obj:
        segment_total += 1
        segment_end = min(segment_start + datetime.timedelta(days=599), end_obj)
        page_num = 1
        while page_num <= max_pages:
            page_total += 1
            payload = {
                "productId": fund_key,
                "startDate": segment_start.strftime('%Y%m%d'),
                "endDate": segment_end.strftime('%Y%m%d'),
                "pageNum": page_num,
                "pageSize": page_size,
            }
            try:
                response = my_fund.session.post(
                    api_url,
                    params={"_csrf": my_fund._csrf},
                    json=payload,
                    headers=headers,
                    timeout=10,
                    verify=False,
                )
                response_json = response.json()
            except Exception as e:
                logger.warning(f"区间历史净值接口请求失败【{fund_code} {segment_start}~{segment_end} page={page_num}】: {e}")
                break

            if not response_json.get("success"):
                break

            value_list = response_json.get("list", []) or []
            if not value_list:
                break

            for item in value_list:
                nav_date = str(item.get("netValueDate", "")).strip()
                if not nav_date:
                    continue
                try:
                    nav_value = float(item.get("netValue"))
                except (TypeError, ValueError):
                    continue
                if nav_value <= 0:
                    continue
                result[nav_date] = round(nav_value, 4)

            if len(value_list) < page_size:
                break
            page_num += 1

        segment_start = segment_end + datetime.timedelta(days=1)

    elapsed_ms = int((time.perf_counter() - request_started_at) * 1000)
    logger.info(
        f"历史净值请求完成【{fund_code}】: start={start_obj}, end={end_obj}, segments={segment_total}, pages={page_total}, records={len(result)}, page_size={page_size}, elapsed_ms={elapsed_ms}"
    )

    return result


def _build_missing_nav_segments(expected_dates: List[datetime.date], existing_nav_map: dict) -> List[Tuple[datetime.date, datetime.date]]:
    """按缺失交易日构建固定 6 个月补齐片段。"""
    if not expected_dates:
        return []
    existing_dates = {datetime.date.fromisoformat(str(day)) for day in existing_nav_map.keys()}
    missing_dates = sorted([day for day in expected_dates if day not in existing_dates])
    if not missing_dates:
        return []
    return _build_backfill_request_segments(missing_dates, missing_dates[-1])


def _sync_nav_history_for_curve(user_id, fund_code, expected_dates: List[datetime.date]):
    """根据业绩/收益曲线需求补齐本地净值（含新鲜度与缺口段判断）。"""
    if not expected_dates:
        return {}

    expected_dates = sorted({item for item in expected_dates if isinstance(item, datetime.date)})
    if not expected_dates:
        return {}

    start_date = expected_dates[0]
    end_date = expected_dates[-1]

    establishment_date = _get_fund_establishment_date(user_id, fund_code)
    if establishment_date and establishment_date > start_date:
        expected_dates = [day for day in expected_dates if day >= establishment_date]
        if not expected_dates:
            return {}
        start_date = expected_dates[0]
        end_date = expected_dates[-1]

    sync_end = min(end_date, _nav_backfill_effective_end_date())
    if sync_end < start_date:
        return db.get_fund_nav_history_range(fund_code, start_date.isoformat(), end_date.isoformat()) or {}

    expected_dates_sync = [d for d in expected_dates if d <= sync_end]
    if not expected_dates_sync:
        return db.get_fund_nav_history_range(fund_code, start_date.isoformat(), end_date.isoformat()) or {}

    local_nav_map = db.get_fund_nav_history_range(fund_code, start_date.isoformat(), end_date.isoformat()) or {}

    # 新鲜度判断：本地最大净值日 < 预期最大交易日则视为不新鲜，触发增量补齐
    local_max_date = None
    if local_nav_map:
        try:
            local_max_date = max(datetime.date.fromisoformat(day) for day in local_nav_map.keys())
        except Exception:
            local_max_date = None

    need_fresh_backfill = local_max_date is None or local_max_date < sync_end
    missing_segments = _build_missing_nav_segments(expected_dates_sync, local_nav_map)

    if not need_fresh_backfill and not missing_segments:
        return local_nav_map

    request_segments = missing_segments[:]
    if need_fresh_backfill:
        fresh_start = (local_max_date + datetime.timedelta(days=1)) if local_max_date else start_date
        if fresh_start <= sync_end:
            request_segments.append((fresh_start, sync_end))

    # 合并重叠请求片段
    request_segments = sorted(request_segments, key=lambda item: item[0])
    merged_segments: List[Tuple[datetime.date, datetime.date]] = []
    for seg_start, seg_end in request_segments:
        if not merged_segments:
            merged_segments.append((seg_start, seg_end))
            continue
        last_start, last_end = merged_segments[-1]
        if seg_start <= (last_end + datetime.timedelta(days=1)):
            merged_segments[-1] = (last_start, max(last_end, seg_end))
        else:
            merged_segments.append((seg_start, seg_end))

    wrote_count = 0
    for seg_start, seg_end in merged_segments:
        remote_nav_map = _fetch_history_nav_map_by_date_range(user_id, fund_code, seg_start.isoformat(), seg_end.isoformat())
        if not remote_nav_map:
            continue
        for nav_date, nav_value in remote_nav_map.items():
            if db.upsert_fund_nav_history(fund_code, nav_date, nav_value, source='history_api_bulk'):
                wrote_count += 1

    if wrote_count > 0:
        logger.info(
            f"净值补齐完成【{fund_code}】: range={start_date}~{end_date}, wrote={wrote_count}, segments={len(merged_segments)}"
        )

    return db.get_fund_nav_history_range(fund_code, start_date.isoformat(), end_date.isoformat()) or local_nav_map


def _normalize_import_tx_type(tx_type_raw):
    tx_type_text = str(tx_type_raw or '').strip().lower()
    if not tx_type_text:
        return None

    normalized_text = tx_type_text.replace(' ', '')
    buy_keywords = ('buy', '买入', '申购', '购入')
    sell_keywords = ('sell', '卖出', '赎回', '卖')
    dividend_keywords = ('dividend', '分红', '现金分红', '派息', '红利')

    if any(keyword in normalized_text for keyword in buy_keywords):
        return 'buy'
    if any(keyword in normalized_text for keyword in sell_keywords):
        return 'sell'
    if any(keyword in normalized_text for keyword in dividend_keywords):
        return 'dividend'
    return None


def _parse_import_trade_datetime(raw_value):
    if isinstance(raw_value, datetime.datetime):
        return raw_value.replace(microsecond=0)

    if isinstance(raw_value, datetime.date):
        return datetime.datetime.combine(raw_value, datetime.time(0, 0, 0))

    if isinstance(raw_value, (int, float)):
        try:
            excel_base = datetime.datetime(1899, 12, 30)
            days = float(raw_value)
            seconds = int(round((days - int(days)) * 24 * 3600))
            return (excel_base + datetime.timedelta(days=int(days), seconds=seconds)).replace(microsecond=0)
        except Exception:
            return None

    text = str(raw_value or '').strip()
    if not text:
        return None

    parse_formats = [
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%Y/%m/%d %H:%M:%S',
        '%Y/%m/%d %H:%M',
        '%Y-%m-%d',
        '%Y/%m/%d',
    ]
    for fmt in parse_formats:
        try:
            parsed = datetime.datetime.strptime(text, fmt)
            return parsed.replace(microsecond=0)
        except Exception:
            continue

    try:
        return datetime.datetime.fromisoformat(text.replace('T', ' ')).replace(microsecond=0)
    except Exception:
        return None


def _normalize_import_text(value):
    if value is None:
        return ''
    text = str(value)
    text = text.replace('\r', '').replace('\n', '').strip()
    return text


def _resolve_net_value_for_trade_datetime(user_id, fund_code, trade_dt):
    """按交易时间解析用于成交换算的净值：交易日15:00前(不含15:00)取当日，否则取下一交易日净值。"""
    if not isinstance(trade_dt, datetime.datetime):
        return None, None

    base_date = trade_dt.date()
    if trade_dt.time() >= datetime.time(15, 0, 0):
        candidate_date = base_date + datetime.timedelta(days=1)
    else:
        candidate_date = base_date

    for _ in range(15):
        target_date = candidate_date.isoformat()
        net_value = db.get_fund_nav_by_date(fund_code, target_date)
        if net_value is None:
            net_value = _find_net_value_by_date_from_history_api(user_id, fund_code, target_date)
            if net_value is not None and net_value > 0:
                db.upsert_fund_nav_history(fund_code, target_date, net_value, source='history_api')

        if net_value is None:
            net_value = _find_net_value_by_date_from_trend(user_id, fund_code, target_date)
            if net_value is not None and net_value > 0:
                db.upsert_fund_nav_history(fund_code, target_date, net_value, source='trend')

        if net_value is not None and net_value > 0:
            return float(net_value), target_date
        candidate_date = candidate_date + datetime.timedelta(days=1)

    return None, None


def _cache_nav_history_from_curve_data(fund_code, labels, net_values, source='performance_curve'):
    """将业绩曲线中的净值数据回填到本地缓存。"""
    cached_count = 0
    if not fund_code:
        return cached_count

    labels = labels or []
    net_values = net_values or []
    max_len = min(len(labels), len(net_values))
    for idx in range(max_len):
        label_text = str(labels[idx] or '').strip()
        if not label_text:
            continue
        try:
            nav_date = datetime.date.fromisoformat(label_text).isoformat()
        except Exception:
            continue

        nav_value = _safe_float(net_values[idx], None)
        if nav_value is None or nav_value <= 0:
            continue

        if db.upsert_fund_nav_history(fund_code, nav_date, nav_value, source=source):
            cached_count += 1

    return cached_count


def _read_excel_rows_for_transaction_import(file_storage):
    if load_workbook is None:
        return None, '缺少 openpyxl 依赖，无法解析Excel，请安装 openpyxl'

    try:
        workbook = load_workbook(file_storage, data_only=True, read_only=True)
    except Exception as e:
        return None, f'Excel解析失败: {e}'

    sheet = workbook.active
    if sheet is None:
        return None, 'Excel解析失败: 未找到可用工作表'
    rows = list(sheet.iter_rows(values_only=False))
    if not rows:
        return None, 'Excel内容为空'

    header_row = rows[0]
    normalized_headers = [str(getattr(cell, 'value', '') or '').strip().lower() for cell in header_row]
    alias_map = {
        'order_no': {'order_no', '订单号', '委托编号', '订单编号', '交易单号'},
        'fund_code': {'fund_code', '基金代码', '代码', '基金'},
        'tx_type': {'tx_type', '交易类型', '类型', '买卖方向', '买卖'},
        'trade_time': {'trade_time', '交易时间', '成交时间', '确认时间', '日期'},
        'amount': {'amount', '确认金额', '金额', '到账金额', '成交金额'},
        'confirmed_shares': {'confirmed_shares', '确认份额', '份额', '成交份额', '确认数量'},
        'fee': {'fee', '手续费', '费用', '交易手续费'},
    }

    col_index = {}
    for field_name, aliases in alias_map.items():
        for idx, header in enumerate(normalized_headers):
            if header in {str(alias).strip().lower() for alias in aliases}:
                col_index[field_name] = idx
                break

    required_fields = ['order_no', 'fund_code', 'tx_type', 'trade_time', 'amount']
    for field_name in required_fields:
        if field_name not in col_index:
            return None, f'Excel缺少必填列: {field_name}'

    parsed_rows = []
    for row_num, row_cells in enumerate(rows[1:], start=2):
        row_dict: dict[str, Any] = {'row_num': row_num}
        for field_name, idx in col_index.items():
            cell = row_cells[idx] if idx < len(row_cells) else None
            raw_value = getattr(cell, 'value', None) if cell is not None else None

            if field_name == 'fund_code':
                if raw_value is None:
                    row_dict[field_name] = None
                elif isinstance(raw_value, str):
                    row_dict[field_name] = _normalize_import_text(raw_value)
                elif isinstance(raw_value, (int, float)):
                    if isinstance(raw_value, float):
                        if raw_value.is_integer():
                            code_text = str(int(raw_value))
                        else:
                            code_text = _normalize_import_text(raw_value)
                    else:
                        code_text = str(int(raw_value))

                    number_format = str(getattr(cell, 'number_format', '') or '').strip()
                    if number_format and set(number_format) == {'0'}:
                        code_text = code_text.zfill(len(number_format))
                    elif code_text.isdigit() and len(code_text) < 6:
                        code_text = code_text.zfill(6)

                    row_dict[field_name] = code_text
                else:
                    row_dict[field_name] = _normalize_import_text(raw_value)
            else:
                if isinstance(raw_value, str):
                    row_dict[field_name] = _normalize_import_text(raw_value)
                else:
                    row_dict[field_name] = raw_value
        parsed_rows.append(row_dict)

    return parsed_rows, None


def _log_import_share_mismatch(fund_code, trade_dt, amount, tx_type, net_value, confirmed_shares, computed_shares):
    tx_label_map = {
        'buy': '买入',
        'sell': '卖出',
    }
    trade_text = trade_dt.strftime('%Y-%m-%d') if isinstance(trade_dt, datetime.datetime) else str(trade_dt or '-')
    tx_label = tx_label_map.get(str(tx_type or '').lower(), str(tx_type or '-'))
    message = (
        f'{fund_code}基金于{trade_text}日的{float(amount or 0):.2f}元{tx_label}交易，'
        f'净值{float(net_value or 0):.4f}，确认份额{float(confirmed_shares or 0):.2f}与计算份额{float(computed_shares or 0):.2f}不一致'
    )
    logger.warning('{}', message)
    _append_import_detail_log(
        'WARNING',
        message,
        fund_code=fund_code,
        trade_time=trade_text,
        tx_type=tx_label,
        amount=f'{float(amount or 0):.2f}',
        net_value=f'{float(net_value or 0):.4f}',
        confirmed_shares=f'{float(confirmed_shares or 0):.2f}',
        computed_shares=f'{float(computed_shares or 0):.2f}',
    )
    return message

def _get_latest_fund_net_value(user_id, fund_code):
    """获取基金最新净值（用于买入/卖出换算）。"""
    net_value, _nav_date, fund_data = _get_latest_fund_quote(user_id, fund_code)
    return net_value, fund_data


def _parse_tx_datetime(tx_time):
    text = str(tx_time or '').strip()
    if not text:
        return None
    try:
        return datetime.datetime.fromisoformat(text.replace(' ', 'T'))
    except Exception:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(text, fmt)
        except Exception:
            continue
    return None


def _calculate_holding_shares_by_time(user_id, fund_code, up_to_dt=None, fallback_shares=0.0):
    """按交易流水重算可用份额。

    关键点：当 up_to_dt 提供时，只统计该时间点(含)之前的买卖，
    这样卖出校验不会把“未来买入”的份额提前拿来卖。
    """
    transactions = db.get_fund_transactions(user_id, fund_code)
    if not transactions:
        # 没有交易流水时，仅“当前卖出”场景可回退到缓存持仓。
        return float(fallback_shares or 0.0) if up_to_dt is None else 0.0

    holding = 0.0
    share_eps = 1e-8

    for tx in transactions:
        tx_dt = _parse_tx_datetime(tx.get('tx_time'))
        if up_to_dt is not None:
            if tx_dt is None or tx_dt > up_to_dt:
                continue

        tx_type = str(tx.get('tx_type', '')).strip().lower()
        tx_shares = _safe_float(tx.get('shares', 0), 0.0) or 0.0
        if tx_shares <= 0:
            continue

        if tx_type == 'buy':
            holding += tx_shares
        elif tx_type == 'sell':
            holding -= tx_shares

        # 浮点误差与脏数据保护：中间不允许出现负持仓。
        if holding < share_eps:
            holding = 0.0

    return holding


def _safe_float(value: Any, default: Optional[float] = 0.0) -> Optional[float]:
    try:
        return float(value)
    except (TypeError, ValueError):
        if default is None:
            return None
        try:
            return float(default)
        except (TypeError, ValueError):
            return None


def _safe_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        if default is None:
            return None
        try:
            return int(default)
        except (TypeError, ValueError):
            return None


def _quantize_shares_2(value):
    """份额统一按两位小数四舍五入。"""
    try:
        return float(Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
    except Exception:
        return 0.0


def _xnpv(rate, cashflows):
    if rate <= -0.999999:
        return float('inf')
    if not cashflows:
        return 0.0
    t0 = cashflows[0][0]
    total_value = 0.0
    for tx_date, amount in cashflows:
        days = (tx_date - t0).total_seconds() / 86400.0
        total_value += amount / ((1 + rate) ** (days / 365.0))
    return total_value


def _solve_xirr(cashflows):
    if len(cashflows) < 2:
        return None
    has_positive = any(amount > 0 for _, amount in cashflows)
    has_negative = any(amount < 0 for _, amount in cashflows)
    if not (has_positive and has_negative):
        return None

    low, high = -0.9999, 10.0
    try:
        f_low = _xnpv(low, cashflows)
        f_high = _xnpv(high, cashflows)
    except Exception:
        return None

    expand_count = 0
    while f_low * f_high > 0 and expand_count < 20:
        high *= 2
        try:
            f_high = _xnpv(high, cashflows)
        except Exception:
            return None
        expand_count += 1

    if f_low * f_high > 0:
        return None

    for _ in range(100):
        mid = (low + high) / 2.0
        f_mid = _xnpv(mid, cashflows)
        if f_mid is None:
            return None
        if abs(f_mid) < 1e-7:
            return mid
        if f_low * f_mid <= 0:
            high = mid
            f_high = f_mid
        else:
            low = mid
            f_low = f_mid
    return (low + high) / 2.0


def _build_clear_cycles(transactions):
    share_eps = 1e-4
    running_shares = 0.0
    cycle_start_dt = None
    cycle_total_buy = 0.0
    cycle_total_sell = 0.0
    cycle_cashflows = []
    clear_cycles = []

    for tx in transactions:
        tx_type = str(tx.get('tx_type', '')).strip().lower()
        tx_shares = _safe_float(tx.get('shares', 0), 0.0) or 0.0
        tx_amount = _safe_float(tx.get('amount', 0), 0.0) or 0.0
        tx_net_value = _safe_float(tx.get('net_value', 0), 0.0) or 0.0
        tx_fee = _safe_float(tx.get('fee', 0), 0.0) or 0.0
        tx_dt = _parse_tx_datetime(tx.get('tx_time'))
        if tx_dt is None:
            continue

        gross_amount = tx_amount if tx_amount > 0 else (tx_shares * tx_net_value)
        if tx_type == 'sell' and tx_fee > 0:
            gross_amount += tx_fee
        effective_amount = gross_amount

        if tx_type == 'buy' and tx_shares > 0:
            if running_shares <= share_eps and cycle_start_dt is None:
                cycle_start_dt = tx_dt
                cycle_total_buy = 0.0
                cycle_total_sell = 0.0
                cycle_cashflows = []
            running_shares += tx_shares
            cycle_total_buy += effective_amount
            cycle_cashflows.append((tx_dt, -effective_amount))
        elif tx_type == 'sell' and tx_shares > 0:
            if running_shares <= share_eps:
                continue
            sell_shares = min(tx_shares, running_shares)
            if sell_shares <= 0:
                continue
            running_shares -= sell_shares
            scale = (sell_shares / tx_shares) if tx_shares > 0 else 0.0
            proceeds = effective_amount * scale
            cycle_total_sell += proceeds
            cycle_cashflows.append((tx_dt, proceeds))

            if running_shares <= share_eps:
                running_shares = 0.0
                period_start = cycle_start_dt or tx_dt
                period_end = tx_dt
                cycle_profit = cycle_total_sell - cycle_total_buy
                cycle_return_pct = (cycle_profit / cycle_total_buy * 100.0) if cycle_total_buy > 0 else None
                annual_rate = _solve_xirr(cycle_cashflows)
                annual_return_pct = (annual_rate * 100.0) if annual_rate is not None else None
                clear_cycles.append({
                    'clear_tx_id': _safe_int(tx.get('id', 0), 0),
                    'period_start': period_start.strftime('%Y-%m-%d'),
                    'period_end': period_end.strftime('%Y-%m-%d'),
                    'cycle_profit': round(cycle_profit, 2),
                    'cycle_return_pct': round(cycle_return_pct, 2) if cycle_return_pct is not None else None,
                    'annual_return_pct': round(annual_return_pct, 2) if annual_return_pct is not None else None,
                })
                cycle_start_dt = None
                cycle_total_buy = 0.0
                cycle_total_sell = 0.0
                cycle_cashflows = []
        elif tx_type == 'dividend' and tx_amount > 0 and running_shares > share_eps:
            cycle_total_sell += tx_amount
            cycle_cashflows.append((tx_dt, tx_amount))

    return clear_cycles

@app.route('/api/fund/shares', methods=['POST'])
@login_required
def api_fund_shares():
    """更新基金持仓份额"""
    try:
        data = request.json or {}
        code = data.get('code', '').strip()
        shares = data.get('shares', 0)

        if not code:
            return {'success': False, 'message': '请提供基金代码'}

        try:
            shares = float(shares)
            if shares < 0:
                return {'success': False, 'message': '份额不能为负数'}
        except (ValueError, TypeError):
            return {'success': False, 'message': '份额格式错误'}

        user_id = get_current_user_id()
        success = db.update_fund_shares(user_id, code, shares)

        if success:
            fund_map = db.get_user_funds(user_id)
            latest = fund_map.get(code, {})
            latest_shares = float(latest.get('shares', shares) or 0)
            latest_is_hold = bool(latest.get('is_hold', latest_shares > 0))
            return {
                'success': True,
                'message': f'已更新份额: {latest_shares:.2f}',
                'current_shares': latest_shares,
                'current_is_hold': latest_is_hold,
            }
        else:
            return {'success': False, 'message': '更新失败，基金不存在'}

    except Exception as e:
        logger.error(f"更新份额失败: {e}")
        return {'success': False, 'message': f'更新失败: {str(e)}'}


@app.route('/api/fund/buy', methods=['POST'])
@login_required
def api_fund_buy():
    """按金额买入基金，按净值日规则延迟确认份额。"""
    try:
        data = request.json or {}
        code = str(data.get('code', '')).strip()
        amount = data.get('amount', 0)
        if not code:
            return {'success': False, 'message': '请提供基金代码'}

        try:
            amount = float(amount)
        except (TypeError, ValueError):
            return {'success': False, 'message': '买入金额格式错误'}
        if amount <= 0:
            return {'success': False, 'message': '买入金额必须大于0'}

        user_id = get_current_user_id()
        user_funds = db.get_user_funds(user_id)
        if code not in user_funds:
            return {'success': False, 'message': '基金不存在'}

        effective_date = _get_buy_effective_date()
        pending_id = db.add_pending_buy(
            user_id=user_id,
            fund_code=code,
            amount=amount,
            effective_date=effective_date,
        )
        if not pending_id:
            return {'success': False, 'message': '买入失败，待确认记录写入失败'}

        _settle_pending_buys(user_id)

        pending_list = db.get_pending_buys(user_id, code)
        is_pending = any(int(item.get('id', -1)) == int(pending_id) for item in pending_list)
        latest_funds = db.get_user_funds(user_id)
        current_shares = float(latest_funds.get(code, {}).get('shares', 0) or 0)

        today_text = datetime.date.today().isoformat()
        if is_pending:
            if effective_date == today_text:
                message = f'买入已提交：¥{amount:,.2f}，将按{effective_date}净值确认份额（15:00后）'
            else:
                message = f'买入已提交：¥{amount:,.2f}，将按{effective_date}及之后首个净值日确认份额'
        else:
            message = f'买入已确认：¥{amount:,.2f}，份额已按净值入账'

        return {
            'success': True,
            'message': message,
            'current_shares': current_shares,
            'pending': is_pending,
            'effective_date': effective_date,
        }
    except Exception as e:
        logger.error(f"买入失败: {e}")
        return {'success': False, 'message': f'买入失败: {str(e)}'}


@app.route('/api/fund/buy-backfill', methods=['POST'])
@login_required
def api_fund_buy_backfill():
    """补录买入：按指定日期净值与金额生成买入交易。"""
    try:
        data = request.json or {}
        code = str(data.get('code', '')).strip()
        amount = data.get('amount', 0)
        fee = data.get('fee', 0)
        net_value = data.get('net_value', None)
        trade_date = str(data.get('trade_date', '')).strip()

        if not code:
            return {'success': False, 'message': '请提供基金代码'}

        try:
            amount = float(amount)
        except (TypeError, ValueError):
            return {'success': False, 'message': '买入金额格式错误'}
        if amount <= 0:
            return {'success': False, 'message': '买入金额必须大于0'}

        try:
            fee = float(fee)
        except (TypeError, ValueError):
            return {'success': False, 'message': '手续费格式错误'}
        if fee < 0:
            return {'success': False, 'message': '手续费不能为负数'}
        if amount <= fee:
            return {'success': False, 'message': '买入金额需大于手续费'}

        try:
            normalized_date = datetime.date.fromisoformat(trade_date).isoformat()
        except Exception:
            return {'success': False, 'message': '交易日期格式错误，请使用YYYY-MM-DD'}

        user_id = get_current_user_id()

        net_value_missing = (net_value is None) or (str(net_value).strip() == "")
        if net_value_missing:
            net_value = _find_net_value_by_date_from_history_api(user_id, code, normalized_date)
            if net_value is None:
                return {'success': False, 'message': '未查询到该日期净值，请手动输入净值后重试'}
        else:
            try:
                net_value = float(str(net_value).strip())
            except (TypeError, ValueError):
                return {'success': False, 'message': '净值格式错误'}
            if net_value <= 0:
                return {'success': False, 'message': '净值必须大于0'}

        net_buy_amount = float((Decimal(str(amount)) - Decimal(str(fee))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        buy_shares = float((Decimal(str(net_buy_amount)) / Decimal(str(net_value))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        if buy_shares <= 0:
            return {'success': False, 'message': '买入金额过小，折算份额为0'}

        user_funds = db.get_user_funds(user_id)
        if code not in user_funds:
            return {'success': False, 'message': '基金不存在'}

        new_shares = db.update_fund_shares_delta(user_id, code, buy_shares)
        if new_shares is None:
            return {'success': False, 'message': '补录失败，份额更新异常'}

        tx_time = f"{normalized_date} 15:00:00"
        tx_id = db.add_fund_transaction(
            user_id=user_id,
            fund_code=code,
            tx_type='buy',
            amount=amount,
            shares=buy_shares,
            net_value=net_value,
            tx_time=tx_time,
            fee=fee,
        )
        if tx_id is None:
            db.update_fund_shares_delta(user_id, code, -buy_shares)
            return {'success': False, 'message': '补录失败，交易记录写入异常'}

        return {
            'success': True,
            'message': f'补录成功：{normalized_date} 按净值 {net_value:.4f} 买入 ¥{amount:,.2f}（手续费¥{fee:,.2f}，{buy_shares:.2f}份）',
            'current_shares': new_shares,
            'trade_date': normalized_date,
            'shares': buy_shares,
            'fee': fee,
        }
    except Exception as e:
        logger.error(f"补录买入失败: {e}")
        return {'success': False, 'message': f'补录买入失败: {str(e)}'}


@app.route('/api/fund/net-value-by-date', methods=['GET'])
@login_required
def api_fund_net_value_by_date():
    """按日期获取基金净值（历史净值接口优先，趋势数据兜底）。"""
    try:
        code = str(request.args.get('code', '')).strip()
        trade_date = str(request.args.get('date', '')).strip()

        if not code:
            return {'success': False, 'message': '请提供基金代码'}

        try:
            normalized_date = datetime.date.fromisoformat(trade_date).isoformat()
        except Exception:
            return {'success': False, 'message': '日期格式错误，请使用YYYY-MM-DD'}

        user_id = get_current_user_id()
        net_value = _find_net_value_by_date_from_history_api(user_id, code, normalized_date)
        if net_value is None:
            net_value = _find_net_value_by_date_from_trend(user_id, code, normalized_date)

        if net_value is None:
            return {
                'success': True,
                'found': False,
                'message': '未找到该日期净值，请手动输入',
                'trade_date': normalized_date,
            }

        return {
            'success': True,
            'found': True,
            'trade_date': normalized_date,
            'net_value': net_value,
        }
    except Exception as e:
        logger.error(f"按日期查询净值失败: {e}")
        return {'success': False, 'message': f'按日期查询净值失败: {str(e)}'}


@app.route('/api/fund/sell-backfill', methods=['POST'])
@login_required
def api_fund_sell_backfill():
    """补录卖出：按份额与指定日期净值生成卖出交易。"""
    try:
        data = request.json or {}
        code = str(data.get('code', '')).strip()
        shares_input = data.get('shares', None)
        fee = data.get('fee', 0)
        net_value = data.get('net_value', None)
        trade_date = str(data.get('trade_date', '')).strip()

        if not code:
            return {'success': False, 'message': '请提供基金代码'}

        shares = None
        if shares_input is not None and str(shares_input).strip() != "":
            try:
                shares = float(shares_input)
            except (TypeError, ValueError):
                return {'success': False, 'message': '卖出份额格式错误'}
            if shares <= 0:
                return {'success': False, 'message': '卖出份额必须大于0'}

        try:
            fee = float(fee)
        except (TypeError, ValueError):
            return {'success': False, 'message': '手续费格式错误'}
        if fee < 0:
            return {'success': False, 'message': '手续费不能为负数'}

        try:
            normalized_date = datetime.date.fromisoformat(trade_date).isoformat()
        except Exception:
            return {'success': False, 'message': '交易日期格式错误，请使用YYYY-MM-DD'}

        user_id = get_current_user_id()
        user_funds = db.get_user_funds(user_id)
        if code not in user_funds:
            return {'success': False, 'message': '基金不存在'}

        # 补录卖出必须按“交易时点之前”可用份额校验，不能用当前总持仓。
        trade_dt = datetime.datetime.combine(
            datetime.date.fromisoformat(normalized_date),
            datetime.time(15, 0, 0),
        )
        current_holding = _calculate_holding_shares_by_time(user_id, code, up_to_dt=trade_dt)

        net_value_missing = (net_value is None) or (str(net_value).strip() == "")
        if net_value_missing:
            net_value = _find_net_value_by_date_from_history_api(user_id, code, normalized_date)
            if net_value is None:
                return {'success': False, 'message': '未查询到该日期净值，请手动输入净值后重试'}
        else:
            try:
                net_value = float(str(net_value).strip())
            except (TypeError, ValueError):
                return {'success': False, 'message': '净值格式错误'}
            if net_value <= 0:
                return {'success': False, 'message': '净值必须大于0'}

        if shares is None:
            return {'success': False, 'message': '请提供卖出份额'}

        gross_sell_amount = float((Decimal(str(shares)) * Decimal(str(net_value))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        if fee > gross_sell_amount:
            return {'success': False, 'message': '手续费不能大于卖出总额'}
        sell_amount = float((Decimal(str(gross_sell_amount)) - Decimal(str(fee))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

        if shares is None or shares <= 0:
            return {'success': False, 'message': '卖出金额过小，折算份额为0'}
        if sell_amount <= 0:
            return {'success': False, 'message': '到账金额必须大于0，请检查份额和手续费'}

        request_shares_2 = _quantize_shares_2(shares)
        available_shares_2 = _quantize_shares_2(current_holding)
        if request_shares_2 > available_shares_2:
            return {
                'success': False,
                'message': (
                    f'卖出份额超过交易时点可用持仓（可用{available_shares_2:.2f}份，'
                    f'本次{request_shares_2:.2f}份；精确可用{current_holding:.4f}，精确本次{shares:.4f}）'
                )
            }

        # 卖出执行按两位份额进行，避免展示可卖但delta扣减出现微小负数。
        shares = request_shares_2
        recalculated = db.recalculate_fund_shares_from_transactions(user_id, code)
        summary_shares = float((recalculated or {}).get('current_shares', user_funds.get(code, {}).get('shares', 0)) or 0)
        if shares > _quantize_shares_2(summary_shares):
            return {
                'success': False,
                'message': (
                    f'卖出份额超过当前总持仓（当前{summary_shares:.2f}份，'
                    f'本次{shares:.2f}份）'
                )
            }
        if code not in user_funds:
            user_funds[code] = {}
        user_funds[code]['shares'] = summary_shares

        new_shares = db.update_fund_shares_delta(user_id, code, -shares)
        if new_shares is None:
            return {'success': False, 'message': '补录卖出失败，份额更新异常'}

        tx_time = f"{normalized_date} 15:00:00"
        tx_id = db.add_fund_transaction(
            user_id=user_id,
            fund_code=code,
            tx_type='sell',
            amount=sell_amount,
            shares=shares,
            net_value=net_value,
            tx_time=tx_time,
            fee=fee,
        )
        if tx_id is None:
            db.update_fund_shares_delta(user_id, code, shares)
            return {'success': False, 'message': '补录卖出失败，交易记录写入异常'}

        return {
            'success': True,
            'message': f'补录卖出成功：{normalized_date} 按净值 {net_value:.4f} 卖出 {shares:.2f}份（到账¥{sell_amount:,.2f}，手续费¥{fee:,.2f}）',
            'current_shares': new_shares,
            'trade_date': normalized_date,
            'shares': shares,
            'amount': sell_amount,
            'fee': fee,
        }
    except Exception as e:
        logger.error(f"补录卖出失败: {e}")
        return {'success': False, 'message': f'补录卖出失败: {str(e)}'}


@app.route('/api/fund/dividend-backfill', methods=['POST'])
@login_required
def api_fund_dividend_backfill():
    """补录分红：按指定日期记录现金分红交易，不变更份额。"""
    try:
        data = request.json or {}
        code = str(data.get('code', '')).strip()
        amount = data.get('amount', 0)
        net_value = data.get('net_value', None)
        trade_date = str(data.get('trade_date', '')).strip()

        if not code:
            return {'success': False, 'message': '请提供基金代码'}

        try:
            amount = float(amount)
        except (TypeError, ValueError):
            return {'success': False, 'message': '分红金额格式错误'}
        if amount <= 0:
            return {'success': False, 'message': '分红金额必须大于0'}

        try:
            normalized_date = datetime.date.fromisoformat(trade_date).isoformat()
        except Exception:
            return {'success': False, 'message': '交易日期格式错误，请使用YYYY-MM-DD'}

        normalized_net_value = None
        if net_value is not None and str(net_value).strip() != '':
            try:
                normalized_net_value = float(str(net_value).strip())
            except (TypeError, ValueError):
                return {'success': False, 'message': '净值格式错误'}
            if normalized_net_value <= 0:
                return {'success': False, 'message': '净值必须大于0'}

        user_id = get_current_user_id()
        user_funds = db.get_user_funds(user_id)
        if code not in user_funds:
            return {'success': False, 'message': '基金不存在'}

        current_shares = float(user_funds.get(code, {}).get('shares', 0) or 0)
        tx_time = f"{normalized_date} 15:00:00"
        tx_id = db.add_fund_transaction(
            user_id=user_id,
            fund_code=code,
            tx_type='dividend',
            amount=amount,
            shares=0,
            net_value=normalized_net_value,
            tx_time=tx_time,
            fee=0,
        )
        if tx_id is None:
            return {'success': False, 'message': '补录分红失败，交易记录写入异常'}

        net_value_desc = f'（参考净值 {normalized_net_value:.4f}）' if normalized_net_value is not None else ''
        return {
            'success': True,
            'message': f'补录分红成功：{normalized_date} 记录现金分红 ¥{amount:,.2f}{net_value_desc}',
            'current_shares': current_shares,
            'trade_date': normalized_date,
            'amount': amount,
            'net_value': normalized_net_value,
        }
    except Exception as e:
        logger.error(f"补录分红失败: {e}")
        return {'success': False, 'message': f'补录分红失败: {str(e)}'}


@app.route('/api/fund/sell', methods=['POST'])
@login_required
def api_fund_sell():
    """按份额卖出基金并记录交易。"""
    try:
        data = request.json or {}
        code = str(data.get('code', '')).strip()
        shares = data.get('shares', 0)
        if not code:
            return {'success': False, 'message': '请提供基金代码'}

        try:
            shares = float(shares)
        except (TypeError, ValueError):
            return {'success': False, 'message': '卖出份额格式错误'}
        if shares <= 0:
            return {'success': False, 'message': '卖出份额必须大于0'}

        user_id = get_current_user_id()
        _settle_pending_buys(user_id)
        user_funds = db.get_user_funds(user_id)
        if code not in user_funds:
            return {'success': False, 'message': '基金不存在'}

        # 当前卖出也按交易流水重算，避免缓存持仓与明细短暂不一致造成误判。
        current_holding = _calculate_holding_shares_by_time(
            user_id,
            code,
            up_to_dt=None,
            fallback_shares=float(user_funds[code].get('shares', 0) or 0),
        )
        request_shares_2 = _quantize_shares_2(shares)
        available_shares_2 = _quantize_shares_2(current_holding)
        if request_shares_2 > available_shares_2:
            return {
                'success': False,
                'message': (
                    f'卖出份额超过当前持仓（可用{available_shares_2:.2f}份，'
                    f'本次{request_shares_2:.2f}份；精确可用{current_holding:.4f}，精确本次{shares:.4f}）'
                )
            }

        # 卖出执行按两位份额进行，避免展示可卖但delta扣减出现微小负数。
        shares = request_shares_2
        summary_shares = float(user_funds.get(code, {}).get('shares', 0) or 0)
        if abs(summary_shares - available_shares_2) > 1e-8:
            db.update_fund_shares(user_id, code, available_shares_2)
            if code not in user_funds:
                user_funds[code] = {}
            user_funds[code]['shares'] = available_shares_2

        net_value, _fund_data = _get_latest_fund_net_value(user_id, code)
        if not net_value or net_value <= 0:
            return {'success': False, 'message': '无法获取基金净值，暂无法卖出'}

        sell_amount = shares * net_value
        new_shares = db.update_fund_shares_delta(user_id, code, -shares)
        if new_shares is None:
            return {'success': False, 'message': '卖出失败，份额更新异常'}

        db.add_fund_transaction(
            user_id=user_id,
            fund_code=code,
            tx_type='sell',
            amount=sell_amount,
            shares=shares,
            net_value=net_value,
            fee=0,
        )

        return {
            'success': True,
            'message': f'卖出成功：{shares:.2f} 份，约 ¥{sell_amount:,.2f}',
            'current_shares': new_shares,
            'net_value': net_value,
        }
    except Exception as e:
        logger.error(f"卖出失败: {e}")
        return {'success': False, 'message': f'卖出失败: {str(e)}'}


@app.route('/api/fund/transactions', methods=['GET'])
@login_required
def api_fund_transactions():
    """获取单只基金交易记录。"""
    try:
        code = str(request.args.get('code', '')).strip()
        if not code:
            return {'success': False, 'message': '请提供基金代码'}

        user_id = get_current_user_id()

        user_funds = db.get_user_funds(user_id)
        if code not in user_funds:
            return {'success': False, 'message': '基金不存在'}

        rows = db.get_fund_transactions(user_id, code)
        share_eps = 1e-6
        running_shares = 0.0
        running_cost = 0.0
        cycle_start_dt = None
        cycle_total_buy = 0.0
        cycle_total_sell = 0.0
        cycle_total_dividend = 0.0
        transactions = []

        def _parse_tx_datetime(tx_time_text):
            tx_text = str(tx_time_text or '').strip()
            if not tx_text:
                return None
            try:
                return datetime.datetime.fromisoformat(tx_text.replace(' ', 'T'))
            except Exception:
                pass
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
                try:
                    return datetime.datetime.strptime(tx_text, fmt)
                except Exception:
                    continue
            return None

        for row in rows:
            tx_type = str(row.get('tx_type', '') or '').lower()
            tx_amount = float(row.get('amount', 0) or 0)
            tx_shares = float(row.get('shares', 0) or 0)
            tx_net_value = float(row.get('net_value', 0) or 0) if row.get('net_value') is not None else 0.0
            tx_time_text = str(row.get('tx_time', '') or '')
            tx_dt = _parse_tx_datetime(tx_time_text)
            liquidation_gain = None
            liquidation_return = None
            holding_days = None

            if tx_type == 'buy' and tx_shares > 0:
                if running_shares <= share_eps:
                    cycle_start_dt = tx_dt
                    cycle_total_buy = 0.0
                    cycle_total_sell = 0.0
                    cycle_total_dividend = 0.0
                buy_cost = tx_amount if tx_amount > 0 else tx_shares * tx_net_value
                running_shares += tx_shares
                running_cost += buy_cost
                cycle_total_buy += buy_cost
            elif tx_type == 'sell' and tx_shares > 0 and running_shares > share_eps:
                avg_cost_before = (running_cost / running_shares) if running_shares > share_eps else 0.0
                sell_shares = min(tx_shares, running_shares)
                running_cost -= sell_shares * avg_cost_before
                running_shares -= sell_shares
                cycle_total_sell += tx_amount
                if running_shares <= share_eps:
                    liquidation_gain = cycle_total_sell + cycle_total_dividend - cycle_total_buy
                    liquidation_return = (liquidation_gain / cycle_total_buy * 100) if cycle_total_buy > share_eps else None
                    if cycle_start_dt and tx_dt:
                        holding_days = max((tx_dt.date() - cycle_start_dt.date()).days, 0)
            elif tx_type == 'dividend' and tx_amount > 0 and running_shares > share_eps:
                cycle_total_dividend += tx_amount

            if running_shares <= share_eps:
                running_shares = 0.0
                running_cost = 0.0
                if tx_type == 'sell':
                    cycle_start_dt = None
                    cycle_total_buy = 0.0
                    cycle_total_sell = 0.0
                    cycle_total_dividend = 0.0

            avg_cost_after = (running_cost / running_shares) if running_shares > share_eps else None

            transactions.append({
                'id': int(row.get('id', 0) or 0),
                'order_no': str(row.get('order_no', '') or ''),
                'fund_code': code,
                'tx_type': tx_type,
                'amount': tx_amount,
                'shares': tx_shares,
                'net_value': float(row.get('net_value', 0) or 0) if row.get('net_value') is not None else None,
                'fee': float(row.get('fee', 0) or 0),
                'tx_time': tx_time_text,
                'avg_cost_after': avg_cost_after,
                'holding_shares_after': running_shares,
                'liquidation_gain': liquidation_gain,
                'liquidation_return': liquidation_return,
                'holding_days': holding_days,
            })

        transactions = transactions[::-1]

        return {
            'success': True,
            'fund_code': code,
            'transactions': transactions,
        }
    except Exception as e:
        logger.error(f"查询交易记录失败: {e}")
        return {'success': False, 'message': f'查询交易记录失败: {str(e)}'}


@app.route('/api/fund/transactions/import', methods=['POST'])
@login_required
def api_fund_transactions_import():
    """从Excel导入交易记录（异步任务）。"""
    try:
        _cleanup_finished_import_jobs()

        if 'file' not in request.files:
            return {'success': False, 'message': '未找到上传文件'}

        file = request.files['file']
        if file.filename == '':
            return {'success': False, 'message': '未选择文件'}

        lower_name = str(file.filename or '').lower()
        if not lower_name.endswith('.xlsx'):
            return {'success': False, 'message': '仅支持 .xlsx 格式Excel文件'}

        file_bytes = file.read()
        if not file_bytes:
            return {'success': False, 'message': '上传文件为空'}

        user_id = get_current_user_id()
        job_id = uuid.uuid4().hex
        _set_import_job_state(
            job_id,
            success=None,
            done=False,
            status='queued',
            title='交易记录导入中',
            detail='文件上传完成，等待服务器处理...',
            percent=0,
            processed_count=0,
            total_count=0,
            imported_count=0,
            failed_count=0,
            duplicate_count=0,
            warning_messages=[],
            failed_rows=[],
            message='导入任务已创建',
        )

        worker = threading.Thread(
            target=_run_transaction_import_job,
            args=(job_id, user_id, file_bytes),
            daemon=True,
        )
        worker.start()

        return {
            'success': True,
            'accepted': True,
            'job_id': job_id,
            'message': '文件上传成功，已开始导入',
        }
    except Exception as e:
        logger.error(f"导入交易记录失败: {e}")
        return {'success': False, 'message': f'导入交易记录失败: {str(e)}'}


def _run_transaction_import_job(job_id, user_id, file_bytes):
    try:
        _set_import_job_state(job_id, status='parsing', detail='正在解析Excel文件...', percent=5)
        parsed_rows, read_error = _read_excel_rows_for_transaction_import(io.BytesIO(file_bytes))
        if read_error:
            _append_import_detail_log('ERROR', f'Excel解析失败: {read_error}', job_id=job_id)
            _set_import_job_state(
                job_id,
                success=False,
                done=True,
                status='failed',
                detail='Excel解析失败',
                percent=100,
                message=read_error,
                failed_count=1,
            )
            return

        if not parsed_rows:
            _append_import_detail_log('ERROR', 'Excel中无可导入数据', job_id=job_id)
            _set_import_job_state(
                job_id,
                success=False,
                done=True,
                status='failed',
                detail='Excel中无可导入数据',
                percent=100,
                message='Excel中无可导入数据',
                failed_count=1,
            )
            return

        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=user_id, db=db)
        user_funds = db.get_user_funds(user_id)

        imported_count = 0
        duplicate_count = 0
        failed_rows = []
        warning_messages = []

        preprocessed_rows = []
        for row in parsed_rows:
            parsed_trade_dt = _parse_import_trade_datetime(row.get('trade_time'))
            preprocessed = dict(row)
            preprocessed['parsed_trade_dt'] = parsed_trade_dt
            preprocessed_rows.append(preprocessed)

        # 先按交易时间从旧到新处理；无法解析时间的记录放到末尾，再按原行号稳定排序
        sorted_rows = sorted(
            preprocessed_rows,
            key=lambda row: (
                row.get('parsed_trade_dt') or datetime.datetime.max,
                int(row.get('row_num', 0) or 0),
            )
        )
        total_count = len(sorted_rows)

        # 缓存：避免同基金同起算日重复查询净值接口
        nav_cache = {}
        # 缓存：避免同一订单号反复查库
        order_exists_cache = {}
        # 缓存：同一导入文件内订单号重复
        seen_order_nos_in_file = set()

        def _get_cached_trade_nav(fund_code, trade_dt):
            if not isinstance(trade_dt, datetime.datetime):
                return None, None
            start_date = trade_dt.date() if trade_dt.time() < datetime.time(15, 0, 0) else (trade_dt.date() + datetime.timedelta(days=1))
            cache_key = (fund_code, start_date.isoformat())
            if cache_key in nav_cache:
                return nav_cache[cache_key]
            resolved = _resolve_net_value_for_trade_datetime(user_id, fund_code, trade_dt)
            nav_cache[cache_key] = resolved
            return resolved

        _set_import_job_state(
            job_id,
            status='processing',
            detail=f'服务器处理中：已处理 0/{total_count}',
            total_count=total_count,
            percent=10,
        )

        for index, item in enumerate(sorted_rows, start=1):
            row_num = int(item.get('row_num', 0) or 0)
            row_diag = {}
            try:
                order_no = _normalize_import_text(item.get('order_no', ''))
                code = _normalize_import_text(item.get('fund_code', ''))
                tx_type = _normalize_import_tx_type(item.get('tx_type'))
                fee = _safe_float(item.get('fee', 0), 0.0)
                trade_dt = item.get('parsed_trade_dt')

                amount_raw = item.get('amount', None)
                amount = _safe_float(amount_raw, None)
                confirmed_shares_raw = item.get('confirmed_shares', None)
                confirmed_shares = _safe_float(confirmed_shares_raw, None)

                if not order_no:
                    raise ValueError('订单号不能为空')

                if order_no in seen_order_nos_in_file:
                    duplicate_count += 1
                    continue

                exists_in_db = order_exists_cache.get(order_no)
                if exists_in_db is None:
                    exists_in_db = db.exists_transaction_order_no(order_no)
                    order_exists_cache[order_no] = exists_in_db
                if exists_in_db:
                    duplicate_count += 1
                    continue

                seen_order_nos_in_file.add(order_no)

                if not code:
                    raise ValueError('基金代码不能为空')
                if tx_type not in ('buy', 'sell', 'dividend'):
                    raise ValueError('交易类型仅支持买入/卖出/分红')
                if trade_dt is None:
                    raise ValueError('交易时间格式错误')
                if amount is None or amount <= 0:
                    raise ValueError('确认金额必须大于0')
                if fee is None or fee < 0:
                    raise ValueError('手续费必须大于等于0')

                if code not in user_funds:
                    my_fund.add_code(code)
                    user_funds = db.get_user_funds(user_id)
                if code not in user_funds:
                    raise ValueError(f'基金代码无效或添加失败: {code}')

                db_snapshot_shares = float(user_funds.get(code, {}).get('shares', 0) or 0)
                row_diag['db_snapshot_shares'] = round(db_snapshot_shares, 6)

                net_value = None
                if tx_type in ('buy', 'sell'):
                    net_value, _nav_date = _get_cached_trade_nav(code, trade_dt)
                    if net_value is None or net_value <= 0:
                        raise ValueError('未查询到可用净值')
                    row_diag['net_value'] = round(float(net_value), 6)

                confirm_amount = float(Decimal(str(amount)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                fee = float(Decimal(str(fee)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                row_diag['confirm_amount'] = round(confirm_amount, 2)
                row_diag['fee'] = round(fee, 2)

                if tx_type == 'buy':
                    buy_base_amount = float((Decimal(str(confirm_amount)) - Decimal(str(fee))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    if buy_base_amount <= 0:
                        raise ValueError('买入确认金额必须大于手续费')
                    row_diag['buy_base_amount'] = round(buy_base_amount, 2)
                    computed_shares = float((Decimal(str(buy_base_amount)) / Decimal(str(net_value))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    if computed_shares <= 0:
                        raise ValueError('买入确认份额计算结果为0')
                    row_diag['computed_shares'] = round(computed_shares, 6)
                    shares = computed_shares
                    if confirmed_shares is not None and confirmed_shares > 0:
                        confirmed_shares = _quantize_shares_2(confirmed_shares)
                        row_diag['confirmed_shares'] = confirmed_shares
                        if abs(confirmed_shares - computed_shares) > 1e-8:
                            warning_messages.append(
                                _log_import_share_mismatch(code, trade_dt, confirm_amount, tx_type, net_value, confirmed_shares, computed_shares)
                            )
                        shares = confirmed_shares
                    new_shares = db.update_fund_shares_delta(user_id, code, shares)
                    if new_shares is None:
                        db_actual_shares = float(db.get_user_funds(user_id).get(code, {}).get('shares', 0) or 0)
                        raise ValueError(
                            f'买入份额更新失败（数据库持仓{db_actual_shares:.2f}, 本次买入{shares:.2f}, '
                            f'确认金额{confirm_amount:.2f}, 手续费{fee:.2f}, 净值{net_value:.4f}）'
                        )
                elif tx_type == 'sell':
                    sell_gross_amount = float((Decimal(str(confirm_amount)) + Decimal(str(fee))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    if sell_gross_amount <= 0:
                        raise ValueError('卖出金额与手续费之和必须大于0')
                    row_diag['sell_gross_amount'] = round(sell_gross_amount, 2)
                    computed_shares = float((Decimal(str(sell_gross_amount)) / Decimal(str(net_value))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    if computed_shares <= 0:
                        raise ValueError('卖出确认份额计算结果为0')
                    row_diag['computed_shares'] = round(computed_shares, 6)
                    shares = computed_shares
                    if confirmed_shares is not None and confirmed_shares > 0:
                        confirmed_shares = _quantize_shares_2(confirmed_shares)
                        row_diag['confirmed_shares'] = confirmed_shares
                        if abs(confirmed_shares - computed_shares) > 1e-8:
                            warning_messages.append(
                                _log_import_share_mismatch(code, trade_dt, confirm_amount, tx_type, net_value, confirmed_shares, computed_shares)
                            )
                        shares = confirmed_shares

                    # 导入时按交易时间顺序校验：只能卖掉该时点之前已买入的份额。
                    current_holding = _calculate_holding_shares_by_time(user_id, code, up_to_dt=trade_dt)
                    row_diag['timeline_holding'] = round(current_holding, 6)
                    request_shares_2 = _quantize_shares_2(shares)
                    available_shares_2 = _quantize_shares_2(current_holding)
                    row_diag['request_shares_2'] = request_shares_2
                    row_diag['available_shares_2'] = available_shares_2
                    if request_shares_2 > available_shares_2:
                        raise ValueError(
                            f'卖出份额超过交易时点可用持仓（可用{available_shares_2:.2f}, 本次{request_shares_2:.2f}, '
                            f'确认金额{confirm_amount:.2f}, 手续费{fee:.2f}, 净值{net_value:.4f}）'
                        )

                    shares = request_shares_2
                    row_diag['executed_shares_2'] = shares

                    # 先按“全量交易流水”回算当前总持仓，再做增量卖出。
                    # 注意不能回写为交易时点可用份额，否则会覆盖掉后续买入导致当前持仓异常。
                    recalculated = db.recalculate_fund_shares_from_transactions(user_id, code)
                    current_total_shares = float((recalculated or {}).get('current_shares', db_snapshot_shares) or 0)
                    row_diag['current_total_shares'] = round(current_total_shares, 6)
                    if shares > _quantize_shares_2(current_total_shares):
                        raise ValueError(
                            f'卖出份额超过当前总持仓（当前{current_total_shares:.2f}, 本次{shares:.2f}）'
                        )
                    if code not in user_funds:
                        user_funds[code] = {}
                    user_funds[code]['shares'] = float(current_total_shares)

                    new_shares = db.update_fund_shares_delta(user_id, code, -shares)
                    if new_shares is None:
                        db_actual_shares = float(db.get_user_funds(user_id).get(code, {}).get('shares', 0) or 0)
                        raise ValueError(
                            f'卖出份额更新失败（数据库持仓{db_actual_shares:.2f}, 交易时点可用{current_holding:.2f}, '
                            f'本次卖出{shares:.2f}, 确认金额{confirm_amount:.2f}, 手续费{fee:.2f}, 净值{net_value:.4f}）'
                        )
                else:
                    # 分红：只记录现金流，不增加份额。
                    shares = 0.0
                    new_shares = float(user_funds.get(code, {}).get('shares', 0) or 0)
                    row_diag['dividend_amount'] = round(confirm_amount, 2)

                tx_time = trade_dt.strftime('%Y-%m-%d %H:%M:%S')
                tx_id = db.add_fund_transaction(
                    user_id=user_id,
                    fund_code=code,
                    tx_type=tx_type,
                    amount=confirm_amount,
                    shares=shares,
                    net_value=net_value,
                    tx_time=tx_time,
                    fee=fee,
                    order_no=order_no,
                )
                if tx_id is None:
                    if tx_type == 'buy':
                        db.update_fund_shares_delta(user_id, code, -shares)
                    elif tx_type == 'sell':
                        db.update_fund_shares_delta(user_id, code, shares)
                    raise ValueError('交易记录写入失败')

                imported_count += 1
                if code not in user_funds:
                    user_funds[code] = {}
                user_funds[code]['shares'] = float(new_shares or 0)
                order_exists_cache[order_no] = True
            except Exception as row_error:
                row_code = _normalize_import_text(item.get('fund_code', '')) or '-'
                row_order_no = _normalize_import_text(item.get('order_no', '')) or '-'
                row_tx_type = _normalize_import_text(item.get('tx_type', '')) or '-'
                row_amount = _normalize_import_text(item.get('amount', '')) or '-'
                row_fee = _normalize_import_text(item.get('fee', '')) or '-'
                parsed_trade_dt = item.get('parsed_trade_dt')
                if isinstance(parsed_trade_dt, datetime.datetime):
                    row_trade_date = parsed_trade_dt.strftime('%Y-%m-%d')
                else:
                    row_trade_date = _normalize_import_text(item.get('trade_time', '')) or '-'
                logger.warning(
                    '交易导入失败: row={}, fund_code={}, tx_type={}, trade_time={}, amount={}, fee={}, diag={}, reason={}',
                    row_num,
                    row_code,
                    row_tx_type,
                    row_trade_date,
                    row_amount,
                    row_fee,
                    json.dumps(row_diag, ensure_ascii=False),
                    str(row_error),
                )
                _append_import_detail_log(
                    'ERROR',
                    f'交易导入失败: 基金{row_code} 日期{row_trade_date}：{str(row_error)}',
                    row=row_num,
                    fund_code=row_code,
                    order_no=row_order_no,
                    tx_type=row_tx_type,
                    trade_time=row_trade_date,
                    amount=row_amount,
                    fee=row_fee,
                    diag=row_diag,
                )
                failed_rows.append({
                    'row': row_num,
                    'reason': f"基金{row_code} 日期{row_trade_date}：{str(row_error)}",
                })
            finally:
                if index % 20 == 0 or index == total_count:
                    process_percent = 10 + int((index / max(total_count, 1)) * 85)
                    _set_import_job_state(
                        job_id,
                        status='processing',
                        detail=f'服务器处理中：已处理 {index}/{total_count}',
                        percent=min(process_percent, 95),
                        processed_count=index,
                        total_count=total_count,
                        imported_count=imported_count,
                        failed_count=len(failed_rows),
                        duplicate_count=duplicate_count,
                        warning_messages=warning_messages[-20:],
                    )

        failed_count = len(failed_rows)
        success = imported_count > 0
        message = (
            f'导入完成：成功 {imported_count} 条，失败 {failed_count} 条，重复订单 {duplicate_count} 条'
            if success else
            f'导入失败，共 {failed_count} 条错误，重复订单 {duplicate_count} 条'
        )
        _set_import_job_state(
            job_id,
            success=success,
            done=True,
            status='completed' if success else 'failed',
            detail=f'服务器处理中：已处理 {total_count}/{total_count}',
            percent=100,
            processed_count=total_count,
            total_count=total_count,
            imported_count=imported_count,
            failed_count=failed_count,
            duplicate_count=duplicate_count,
            warning_messages=warning_messages[-20:],
            failed_rows=failed_rows,
            message=message,
        )
    except Exception as e:
        logger.error(f"交易记录导入后台任务失败: {e}")
        _append_import_detail_log('ERROR', f'交易记录导入后台任务失败: {e}', job_id=job_id)
        state = _get_import_job_state(job_id) or {}
        total_count = int(state.get('total_count', 0) or 0)
        processed_count = int(state.get('processed_count', 0) or 0)
        _set_import_job_state(
            job_id,
            success=False,
            done=True,
            status='failed',
            detail=f'服务器处理中：已处理 {processed_count}/{total_count}' if total_count else '服务器处理失败',
            percent=100,
            message=f'导入交易记录失败: {str(e)}',
        )


@app.route('/api/fund/transactions/import-progress', methods=['GET'])
@login_required
def api_fund_transactions_import_progress():
    job_id = str(request.args.get('job_id', '') or '').strip()
    if not job_id:
        return {'success': False, 'message': '缺少 job_id'}

    state = _get_import_job_state(job_id)
    if not state:
        return {'success': False, 'message': '导入任务不存在或已过期'}

    return {'success': True, 'job': state}


@app.route('/api/fund/transaction/update', methods=['POST'])
@login_required
def api_fund_transaction_update():
    """更新交易记录并重算持仓份额。"""
    try:
        data = request.json or {}
        code = str(data.get('code', '')).strip()
        tx_id_raw = data.get('transaction_id', None)
        tx_type = str(data.get('tx_type', '')).strip().lower()
        amount_raw = data.get('amount', 0)
        shares_raw = data.get('shares', 0)
        net_value_raw = data.get('net_value', 0)
        fee_raw = data.get('fee', 0)
        tx_time_raw = str(data.get('tx_time', '')).strip()

        if not code:
            return {'success': False, 'message': '请提供基金代码'}

        try:
            tx_id = int(str(tx_id_raw).strip())
        except (TypeError, ValueError):
            return {'success': False, 'message': '交易ID格式错误'}
        if tx_id <= 0:
            return {'success': False, 'message': '交易ID无效'}

        if tx_type not in ('buy', 'sell', 'dividend'):
            return {'success': False, 'message': '交易类型必须为买入、卖出或分红'}

        try:
            amount = float(amount_raw)
            fee = float(fee_raw)
        except (TypeError, ValueError):
            return {'success': False, 'message': '金额/手续费格式错误'}

        shares = 0.0
        net_value = None
        if tx_type in ('buy', 'sell'):
            try:
                shares = float(shares_raw)
                net_value = float(net_value_raw)
            except (TypeError, ValueError):
                return {'success': False, 'message': '份额/净值格式错误'}
            shares = float(Decimal(str(shares)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
            if shares <= 0 or net_value <= 0:
                return {'success': False, 'message': '份额、净值都必须大于0'}

        if amount <= 0:
            return {'success': False, 'message': '金额必须大于0'}
        if fee < 0:
            return {'success': False, 'message': '手续费不能为负数'}

        if not tx_time_raw:
            return {'success': False, 'message': '交易时间不能为空'}
        parsed_dt = None
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M', '%Y-%m-%dT%H:%M:%S'):
            try:
                parsed_dt = datetime.datetime.strptime(tx_time_raw, fmt)
                break
            except Exception:
                continue
        if parsed_dt is None:
            return {'success': False, 'message': '交易时间格式错误'}
        tx_time = parsed_dt.strftime('%Y-%m-%d %H:%M:%S')

        user_id = get_current_user_id()
        user_funds = db.get_user_funds(user_id)
        if code not in user_funds:
            return {'success': False, 'message': '基金不存在'}

        result = db.update_fund_transaction_and_recalculate(
            user_id=user_id,
            fund_code=code,
            tx_id=tx_id,
            tx_type=tx_type,
            amount=amount,
            shares=shares,
            net_value=net_value,
            tx_time=tx_time,
            fee=fee,
        )
        if not result:
            return {'success': False, 'message': '更新失败，交易不存在或处理异常'}

        return {
            'success': True,
            'message': '交易记录已更新',
            'current_shares': float(result.get('current_shares', 0) or 0),
            'current_is_hold': bool(result.get('current_is_hold', False)),
        }
    except Exception as e:
        logger.error(f"更新交易记录失败: {e}")
        return {'success': False, 'message': f'更新交易记录失败: {str(e)}'}


@app.route('/api/fund/transaction/delete', methods=['POST'])
@login_required
def api_fund_transaction_delete():
    """删除交易记录并重算持仓份额。"""
    try:
        data = request.json or {}
        code = str(data.get('code', '')).strip()
        tx_id_raw = data.get('transaction_id', None)

        if not code:
            return {'success': False, 'message': '请提供基金代码'}

        try:
            tx_id = int(str(tx_id_raw).strip())
        except (TypeError, ValueError):
            return {'success': False, 'message': '交易ID格式错误'}
        if tx_id <= 0:
            return {'success': False, 'message': '交易ID无效'}

        user_id = get_current_user_id()

        result = db.delete_fund_transaction_and_recalculate(user_id, code, tx_id)
        if not result:
            return {'success': False, 'message': '删除失败，交易不存在或处理异常'}

        deleted_tx = result.get('deleted', {})
        deleted_type_map = {
            'buy': '买入',
            'sell': '卖出',
            'dividend': '分红',
        }
        deleted_type = deleted_type_map.get(str(deleted_tx.get('tx_type', '')).lower(), '交易')
        deleted_shares = float(deleted_tx.get('shares', 0) or 0)

        return {
            'success': True,
            'message': f'已删除{deleted_type}记录（{deleted_shares:.2f}份）',
            'current_shares': float(result.get('current_shares', 0) or 0),
            'current_is_hold': bool(result.get('current_is_hold', False)),
            'deleted_id': int(deleted_tx.get('id', 0) or 0),
        }
    except Exception as e:
        logger.error(f"删除交易记录失败: {e}")
        return {'success': False, 'message': f'删除交易记录失败: {str(e)}'}


@app.route('/api/fund/transactions/clear', methods=['POST'])
@login_required
def api_fund_transactions_clear():
    """清空单只基金的全部交易记录。"""
    try:
        data = request.json or {}
        code = str(data.get('code', '')).strip()
        confirm_text = str(data.get('confirm_text', '')).strip()

        if not code:
            return {'success': False, 'message': '请提供基金代码'}

        expected_confirm = f'清空 {code}'
        if confirm_text != expected_confirm:
            return {'success': False, 'message': f'确认文本不匹配，请输入“{expected_confirm}”'}

        user_id = get_current_user_id()
        user_funds = db.get_user_funds(user_id)
        if code not in user_funds:
            return {'success': False, 'message': '基金不存在'}

        result = db.clear_fund_transactions_and_recalculate(user_id, code)
        if not result:
            return {'success': False, 'message': '清空失败，处理异常'}

        deleted_count = int(result.get('deleted_count', 0) or 0)
        logger.warning('已清空基金交易记录: user_id={}, fund_code={}, deleted_count={}', user_id, code, deleted_count)
        return {
            'success': True,
            'message': f'已清空 {code} 的交易记录，共删除 {deleted_count} 条',
            'deleted_count': deleted_count,
            'current_shares': float(result.get('current_shares', 0) or 0),
            'current_is_hold': bool(result.get('current_is_hold', False)),
        }
    except Exception as e:
        logger.error(f"清空交易记录失败: {e}")
        return {'success': False, 'message': f'清空交易记录失败: {str(e)}'}


@app.route('/api/fund/transactions/clear-all', methods=['POST'])
@login_required
def api_fund_transactions_clear_all():
    """清空当前用户全部基金交易记录。"""
    try:
        data = request.json or {}
        confirm_text = str(data.get('confirm_text', '')).strip()
        expected_confirm = '清空全部交易'
        if confirm_text != expected_confirm:
            return {'success': False, 'message': f'确认文本不匹配，请输入“{expected_confirm}”'}

        user_id = get_current_user_id()
        result = db.clear_all_fund_transactions_and_recalculate(user_id)
        if not result:
            return {'success': False, 'message': '清空失败，处理异常'}

        deleted_count = int(result.get('deleted_count', 0) or 0)
        affected_funds = int(result.get('affected_funds', 0) or 0)
        logger.warning('已清空全部基金交易记录: user_id={}, deleted_count={}, affected_funds={}', user_id, deleted_count, affected_funds)
        return {
            'success': True,
            'message': f'已清空全部交易记录，共删除 {deleted_count} 条，影响 {affected_funds} 只基金',
            'deleted_count': deleted_count,
            'affected_funds': affected_funds,
        }
    except Exception as e:
        logger.error(f"清空全部交易记录失败: {e}")
        return {'success': False, 'message': f'清空全部交易记录失败: {str(e)}'}


@app.route('/api/fund/data', methods=['GET'])
@login_required
def api_fund_data():
    """获取用户的基金数据（用于前端加载份额等信息）"""
    try:
        user_id = get_current_user_id()
        _settle_pending_buys(user_id)
        fund_map = db.get_user_funds(user_id)
        return jsonify(fund_map)
    except Exception as e:
        logger.error(f"获取基金数据失败: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/tab/<tab_id>', methods=['GET'])
@login_required
def api_get_tab_data(tab_id):
    """按需加载单个tab的数据"""
    try:
        user_id = get_current_user_id()
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=user_id, db=db)

        # 定义tab ID到函数的映射
        tab_functions = {
            'kx': my_fund.kx_html,
            'marker': my_fund.marker_html,
            'real_time_gold': my_fund.real_time_gold_html,
            'gold': my_fund.gold_html,
            'seven_A': my_fund.seven_A_html,
            'A': my_fund.A_html,
            'fund': my_fund.fund_html,
            'bk': my_fund.bk_html,
            'select_fund': my_fund.select_fund_html,
        }

        if tab_id not in tab_functions:
            return jsonify({'success': False, 'message': f'未知的tab ID: {tab_id}'}), 404

        # 获取数据
        content = tab_functions[tab_id]()

        # 如果是fund tab，需要增强内容（传递份额数据）
        if tab_id == 'fund':
            user_id = get_current_user_id()
            fund_map = db.get_user_funds(user_id)
            shares_map = {code: data.get('shares', 0) for code, data in fund_map.items()}
            content = enhance_fund_tab_content(content, shares_map)

        return jsonify({'success': True, 'content': content})
    except Exception as e:
        logger.error(f"加载tab {tab_id} 数据失败: {e}")
        return jsonify({'success': False, 'message': f'数据加载失败: {str(e)}'}), 500


# ==================== New API Endpoints for Auto-Refresh ====================

@app.route('/api/timing', methods=['GET'])
@login_required
def api_timing_data():
    """获取上证分时数据"""
    try:
        user_id = get_current_user_id()
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=user_id, db=db)

        # 使用现有的 get_timing_chart_data 方法
        data = my_fund.get_timing_chart_data()
        response_data: dict[str, Any] = dict(data)

        # 添加当前价格和涨跌幅信息（使用原始数据中的正确涨跌幅）
        prices = response_data.get('prices') or []
        if prices:
            response_data['current_price'] = prices[-1]
            change_amounts = response_data.get('change_amounts') or []
            change_pcts = response_data.get('change_pcts') or []
            response_data['change'] = change_amounts[-1] if change_amounts else 0
            response_data['change_pct'] = change_pcts[-1] if change_pcts else 0

        return jsonify({'success': True, 'data': response_data})
    except Exception as e:
        logger.error(f"获取上证分时数据失败: {e}")
        return jsonify({'success': False, 'message': f'数据加载失败: {str(e)}'}), 500


@app.route('/api/news/7x24', methods=['GET'])
@login_required
def api_news_7x24():
    """获取7*24快讯"""
    try:
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=get_current_user_id(), db=db)

        # 获取快讯数据
        result = my_fund.kx(True)

        # 将数据转换为JSON格式
        # kx() 返回的是 list of dicts，需要正确提取字段
        news_items = []
        if result:
            for item in result:
                try:
                    # 提取标题和内容
                    title = item.get('title', '')
                    if not title and 'content' in item and 'items' in item['content']:
                        content_items = item['content'].get('items', [])
                        if content_items and len(content_items) > 0:
                            title = content_items[0].get('data', '')

                    # 提取发布时间
                    publish_time = item.get('publish_time', '')
                    if publish_time:
                        # 转换时间戳为可读格式
                        import datetime
                        try:
                            publish_time = datetime.datetime.fromtimestamp(int(publish_time)).strftime("%H:%M:%S")
                        except:
                            publish_time = ''

                    # 提取评估（利好/利空）
                    evaluate = item.get('evaluate', '')

                    news_items.append({
                        'time': publish_time,
                        'content': title,
                        'source': evaluate if evaluate else ''
                    })
                except Exception as e:
                    logger.debug(f"Error processing news item: {e}")
                    continue

        return jsonify({'success': True, 'data': news_items})
    except Exception as e:
        logger.error(f"获取7*24快讯失败: {e}")
        return jsonify({'success': False, 'message': f'数据加载失败: {str(e)}'}), 500


@app.route('/api/indices/global', methods=['GET'])
@login_required
def api_indices_global():
    """获取全球指数数据"""
    try:
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=get_current_user_id(), db=db)

        # 获取全球指数数据 - 使用正确的方法名
        result = my_fund.get_market_info(True)

        # 将数据转换为JSON格式
        # result 格式: [[名称, 指数, 涨跌幅], ...]
        indices = []
        if result:
            for item in result:
                if len(item) >= 3:
                    # 清理涨跌幅中的颜色代码和%符号
                    change_str = item[2] if item[2] else "0%"
                    change_str = change_str.replace('%', '').replace('\033[1;31m', '').replace('\033[1;32m', '')
                    change = float(change_str) if change_str else 0
                    indices.append({
                        'name': item[0],
                        'value': item[1],
                        'change': change_str + '%',
                        'change_pct': change
                    })

        return jsonify({'success': True, 'data': indices})
    except Exception as e:
        logger.error(f"获取全球指数失败: {e}")
        return jsonify({'success': False, 'message': f'数据加载失败: {str(e)}'}), 500


@app.route('/api/indices/volume', methods=['GET'])
@login_required
def api_indices_volume():
    """获取成交量趋势数据"""
    try:
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=get_current_user_id(), db=db)

        # 使用现有的 get_volume_chart_data 方法
        data = my_fund.get_volume_chart_data()

        return jsonify({'success': True, 'data': data})
    except Exception as e:
        logger.error(f"获取成交量趋势失败: {e}")
        return jsonify({'success': False, 'message': f'数据加载失败: {str(e)}'}), 500


@app.route('/api/gold/real-time', methods=['GET'])
@login_required
def api_gold_realtime():
    """获取实时贵金属价格"""
    try:
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=get_current_user_id(), db=db)

        # 获取实时金价数据
        # real_time_gold 返回 [[...], [...], [...]] 三个贵金属的数据
        # 每个数组有10列: [名称, 最新价, 涨跌额, 涨跌幅, 开盘价, 最高价, 最低价, 昨收价, 更新时间, 单位]
        result = my_fund.real_time_gold(True)

        # 将数据转换为JSON格式，保留所有10列
        gold_data = []
        gold_names = ['中国黄金', '周大福', '周生生']  # 根据API代码 JO_71, JO_92233, JO_92232

        if result and len(result) >= 3:
            # result[0], result[1], result[2] 分别是三种贵金属的数据
            for i, gold_type_data in enumerate(result):
                if len(gold_type_data) >= 4:  # 至少需要前4列
                    gold_data.append({
                        'name': gold_type_data[0] if gold_type_data else gold_names[i],
                        'price': gold_type_data[1] if len(gold_type_data) > 1 else '',
                        'change_amount': gold_type_data[2] if len(gold_type_data) > 2 else '',
                        'change_pct': gold_type_data[3] if len(gold_type_data) > 3 else '',
                        'open_price': gold_type_data[4] if len(gold_type_data) > 4 else '',
                        'high_price': gold_type_data[5] if len(gold_type_data) > 5 else '',
                        'low_price': gold_type_data[6] if len(gold_type_data) > 6 else '',
                        'prev_close': gold_type_data[7] if len(gold_type_data) > 7 else '',
                        'update_time': gold_type_data[8] if len(gold_type_data) > 8 else '',
                        'unit': gold_type_data[9] if len(gold_type_data) > 9 else ''
                    })

        return jsonify({'success': True, 'data': gold_data})
    except Exception as e:
        logger.error(f"获取实时金价失败: {e}")
        return jsonify({'success': False, 'message': f'数据加载失败: {str(e)}'}), 500


@app.route('/api/gold/history', methods=['GET'])
@login_required
def api_gold_history():
    """获取历史金价数据"""
    try:
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=get_current_user_id(), db=db)

        # 获取历史金价数据 (gold 是静态方法，返回 raw data)
        result = my_fund.gold(True)

        # gold 返回格式: [[日期, 中国黄金基础金价, 周大福金价, 中国黄金基础金价涨跌, 周大福金价涨跌], ...]
        gold_history = []
        if result:
            for item in result:
                if len(item) >= 3:
                    gold_history.append({
                        'date': item[0],
                        'china_gold_price': item[1],
                        'chow_tai_fook_price': item[2],
                        'china_gold_change': item[3] if len(item) > 3 else '',
                        'chow_tai_fook_change': item[4] if len(item) > 4 else ''
                    })

        return jsonify({'success': True, 'data': gold_history})
    except Exception as e:
        logger.error(f"获取历史金价失败: {e}")
        return jsonify({'success': False, 'message': f'数据加载失败: {str(e)}'}), 500


@app.route('/api/sectors', methods=['GET'])
@login_required
def api_sectors():
    """获取行业板块数据"""
    try:
        importlib.reload(fund)

        # 获取板块数据 (bk 是静态方法，返回 raw data)
        # 需要从API获取板块代码
        import requests
        url = "https://push2.eastmoney.com/api/qt/clist/get"
        params = {
            "cb": "",
            "fid": "f62",
            "po": "1",
            "pz": "100",
            "pn": "1",
            "np": "1",
            "fltt": "2",
            "invt": "2",
            "ut": "8dec03ba335b81bf4ebdf7b29ec27d15",
            "fs": "m:90 t:2",
            "fields": "f12,f14,f2,f3,f62,f184,f66,f69,f72,f75,f78,f81,f84,f87,f204,f205,f124,f1,f13"
        }
        response = requests.get(url, params=params, timeout=10, verify=False)
        if str(response.json()["data"]):
            data = response.json()["data"]
            sectors = []
            for bk in data["diff"]:
                sectors.append({
                    'code': bk["f12"],  # 板块代码
                    'name': bk["f14"],  # 板块名称
                    'change': str(bk["f3"]) + "%",  # 涨跌幅
                    'main_inflow': str(round(bk["f62"] / 100000000, 2)) + "亿",  # 主力净流入
                    'main_inflow_pct': str(round(bk["f184"], 2)) + "%",  # 主力净流入占比
                    'small_inflow': str(round(bk["f84"] / 100000000, 2)) + "亿",  # 小单净流入
                    'small_inflow_pct': str(round(bk["f87"], 2)) + "%"  # 小单流入占比
                })

            # 按涨跌幅降序排序（与原始 bk() 函数的排序逻辑一致）
            sectors = sorted(
                sectors,
                key=lambda x: float(x['change'].replace("%", "")) if x['main_inflow_pct'] != "N/A" else -99,
                reverse=True
            )
        else:
            sectors = []

        return jsonify({'success': True, 'data': sectors})
    except Exception as e:
        logger.error(f"获取行业板块失败: {e}")
        return jsonify({'success': False, 'message': f'数据加载失败: {str(e)}'}), 500


@app.route('/api/fund/list', methods=['GET'])
@login_required
def api_fund_list():
    """获取基金列表（含份额数据）"""
    try:
        user_id = get_current_user_id()
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=user_id, db=db)

        # 获取用户基金数据
        fund_map = db.get_user_funds(user_id)

        # 将数据转换为JSON格式
        funds = []
        for code, data in fund_map.items():
            fund_info = my_fund.CACHE_MAP.get(code, {})
            funds.append({
                'code': code,
                'name': data.get('fund_name', fund_info.get('name', '')),
                'shares': data.get('shares', 0),
                'is_hold': data.get('is_hold', False),
                'sectors': data.get('sectors', []),
                'net_value': fund_info.get('net_value', 0),
                'day_growth': fund_info.get('day_growth', 0),
                'estimated_growth': fund_info.get('estimated_growth', 0)
            })

        return jsonify({'success': True, 'data': funds})
    except Exception as e:
        logger.error(f"获取基金列表失败: {e}")
        return jsonify({'success': False, 'message': f'数据加载失败: {str(e)}'}), 500


@app.route('/api/sector/<sector_id>', methods=['GET'])
@login_required
def api_sector_funds(sector_id):
    """获取指定板块的基金列表"""
    try:
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=get_current_user_id(), db=db)

        # 获取板块基金数据
        result = my_fund.select_fund(bk_id=sector_id, is_return=True)

        # 将数据转换为JSON格式
        funds = []
        if result:
            for item in result:
                if len(item) >= 5:
                    funds.append({
                        'code': item[0],
                        'name': item[1],
                        'net_value': item[2],
                        'day_growth': item[3],
                        'estimated_growth': item[4] if len(item) > 4 else ''
                    })

        return jsonify({'success': True, 'data': funds})
    except Exception as e:
        logger.error(f"获取板块基金失败: {e}")
        return jsonify({'success': False, 'message': f'数据加载失败: {str(e)}'}), 500


@app.route('/', methods=['GET'])
@login_required
def get_index():
    # 重定向到持仓基金页面
    return redirect('/portfolio')


@app.route('/fund', methods=['GET'])
@login_required
def get_fund():
    # 重定向到持仓基金页面
    return redirect('/portfolio')


@app.route('/market', methods=['GET'])
@login_required
def get_market():
    # old 7*24快讯 route removed, redirect to portfolio
    return redirect('/portfolio')


@app.route('/precious-metals', methods=['GET'])
@login_required
def get_precious_metals():
    """贵金属行情页面"""
    user_id = get_current_user_id()
    importlib.reload(fund)
    my_fund = fund.LanFund(user_id=user_id, db=db)

    # 加载贵金属数据
    precious_metals_data = {}
    try:
        precious_metals_data['real_time'] = my_fund.real_time_gold_html()
        logger.debug("✓ 实时贵金属")
    except Exception as e:
        precious_metals_data['real_time'] = f"<p style='color:#f44336;'>加载失败: {str(e)}</p>"

    try:
        precious_metals_data['history'] = my_fund.gold_html()
        logger.debug("✓ 历史金价")
    except Exception as e:
        precious_metals_data['history'] = f"<p style='color:#f44336;'>加载失败: {str(e)}</p>"

    import src.module_html as module_html
    render_page = getattr(module_html, 'get_precious_metals_page_html', None)
    if callable(render_page):
        html = str(render_page(precious_metals_data, username=get_current_username()))
    else:
        html = str(precious_metals_data.get('real_time', '')) + str(precious_metals_data.get('history', ''))
    return html


@app.route('/market-indices', methods=['GET'])
@login_required
def get_market_indices():
    # market indices page removed; redirect to portfolio
    return redirect('/portfolio')


@app.route('/api/portfolio/fund-table', methods=['GET'])
@login_required
def api_portfolio_fund_table():
    """获取投资组合的基金表格数据（用于刷新）"""
    try:
        user_id = get_current_user_id()
        _settle_pending_buys(user_id)
        importlib.reload(fund)
        my_fund = fund.LanFund(user_id=user_id, db=db)
        fund_table_html = my_fund.fund_html()
        fund_map = db.get_user_funds(user_id)
        shares_map = {code: data.get('shares', 0) for code, data in fund_map.items()}
        fund_table_html = enhance_fund_tab_content(fund_table_html, shares_map)
        return jsonify({
            'success': True,
            'html': fund_table_html
        })
    except Exception as e:
        logger.error(f"获取基金表格失败: {e}")
        return jsonify({
            'success': False,
            'message': f'获取基金表格失败: {str(e)}'
        }), 500


@app.route('/portfolio', methods=['GET'])
@login_required
def get_portfolio():
    """持仓基金页面"""
    add = request.args.get("add")
    delete = request.args.get("delete")
    user_id = get_current_user_id()
    _settle_pending_buys(user_id)
    importlib.reload(fund)
    my_fund = fund.LanFund(user_id=user_id, db=db)
    if add:
        my_fund.add_code(add)
    if delete:
        my_fund.delete_code(delete)

    # 加载基金数据
    try:
        fund_content = my_fund.fund_html()
        # 获取用户份额数据并传递给enhance_fund_tab_content
        fund_map = db.get_user_funds(user_id)
        shares_map = {code: data.get('shares', 0) for code, data in fund_map.items()}
        fund_content = enhance_fund_tab_content(fund_content, shares_map)
    except Exception as e:
        fund_content = f"<p style='color:#f44336;'>数据加载失败: {str(e)}</p>"

    # 获取用户基金列表
    user_funds = db.get_user_funds(user_id)

    # 确定默认显示的基金
    default_fund = None
    fund_chart_data = None
    fund_chart_info = {}

    if user_funds:
        # 1. 检查是否有用户设置的默认基金
        saved_default = db.get_chart_default_fund(user_id)
        if saved_default and saved_default['fund_code'] in user_funds:
            default_fund = saved_default
        # 2. 选择有持仓的基金（预估收益最高的）
        else:
            held_funds = {code: data for code, data in user_funds.items() if data.get('shares', 0) > 0}
            if held_funds:
                # 简化处理：选择第一个有持仓的基金
                first_code = list(held_funds.keys())[0]
                default_fund = {
                    'fund_code': first_code,
                    'fund_key': held_funds[first_code]['fund_key'],
                    'fund_name': held_funds[first_code]['fund_name']
                }
            # 3. 选择自选列表第一个
            else:
                first_code = list(user_funds.keys())[0]
                default_fund = {
                    'fund_code': first_code,
                    'fund_key': user_funds[first_code]['fund_key'],
                    'fund_name': user_funds[first_code]['fund_name']
                }

        # 加载图表数据
        if default_fund:
            fund_chart_data = my_fund.get_fund_chart_data(default_fund['fund_code'], default_fund)

        # 准备基金选择器信息
        for code, data in user_funds.items():
            fund_chart_info[code] = {
                'name': data['fund_name'],
                'is_default': (default_fund and code == default_fund['fund_code'])
            }

    from src.module_html import get_portfolio_page_html
    html = get_portfolio_page_html(
        fund_content=fund_content,
        fund_map=my_fund.CACHE_MAP,
        fund_chart_data=fund_chart_data,
        fund_chart_info=fund_chart_info,
        username=get_current_username()
    )
    return html


@app.route('/api/fund/chart-data')
@login_required
def api_fund_chart_data():
    """获取基金估值趋势图数据"""
    fund_code = request.args.get('code')
    if not fund_code:
        return jsonify({'error': 'Missing fund code'}), 400

    user_id = get_current_user_id()
    user_funds = db.get_user_funds(user_id)

    if fund_code not in user_funds:
        return jsonify({'error': 'Fund not in user list'}), 400

    fund_data = {
        'fund_key': user_funds[fund_code]['fund_key'],
        'fund_name': user_funds[fund_code]['fund_name']
    }

    importlib.reload(fund)
    my_fund = fund.LanFund(user_id=user_id, db=db)
    chart_data = my_fund.get_fund_chart_data(fund_code, fund_data)

    return jsonify({
        'chart_data': chart_data,
        'fund_info': {
            'code': fund_code,
            'name': fund_data['fund_name']
        }
    })


@app.route('/api/fund/performance-chart-data')
@login_required
def api_fund_performance_chart_data():
    """获取基金业绩曲线数据。"""
    fund_code = request.args.get('code')
    date_interval = request.args.get('interval', DEFAULT_PERFORMANCE_CHART_INTERVAL).strip().upper()
    if not fund_code:
        return jsonify({'error': 'Missing fund code'}), 400

    if date_interval not in PERFORMANCE_CHART_INTERVALS:
        return jsonify({'error': 'Invalid interval'}), 400

    user_id = get_current_user_id()
    user_funds = db.get_user_funds(user_id)

    if fund_code not in user_funds:
        return jsonify({'error': 'Fund not in user list'}), 400

    fund_data = {
        'fund_key': user_funds[fund_code]['fund_key'],
        'fund_name': user_funds[fund_code]['fund_name']
    }

    chart_data = _build_local_performance_chart_data(
        user_id=user_id,
        fund_code=fund_code,
        fund_name=fund_data['fund_name'],
        date_interval=date_interval,
        user_funds=user_funds,
    )

    chart_labels = chart_data.get('labels', []) or []
    resolved_label_dates = _resolve_curve_label_dates(chart_labels)
    chart_net_values = chart_data.get('net_values', []) or []

    transactions = db.get_fund_transactions(user_id, fund_code)
    chart_growth = chart_data.get('growth', []) or []
    growth_by_label = {}
    for index, label in enumerate(chart_labels):
        if index >= len(chart_growth):
            continue
        point_date = resolved_label_dates[index] if index < len(resolved_label_dates) else None
        if point_date is not None:
            growth_by_label[point_date.isoformat()] = chart_growth[index]
        else:
            growth_by_label[str(label)] = chart_growth[index]

    clear_cycles = _build_clear_cycles(transactions)
    clear_cycle_map = {item['clear_tx_id']: item for item in clear_cycles}

    trade_markers = []
    for tx in transactions:
        tx_time = str(tx.get('tx_time', '')).strip()
        if not tx_time:
            continue
        tx_date = tx_time.split(' ')[0]
        point_value = growth_by_label.get(tx_date)
        if point_value is None:
            continue
        tx_type = str(tx.get('tx_type', '')).strip().lower()
        tx_id = _safe_int(tx.get('id', 0), 0)
        cycle_info = clear_cycle_map.get(tx_id)
        marker_type = 'clear' if cycle_info else tx_type
        marker_item = {
            'type': tx_type,
            'marker_type': marker_type,
            'x': tx_date,
            'y': point_value,
            'amount': _safe_float(tx.get('amount', 0), 0.0),
            'shares': _safe_float(tx.get('shares', 0), 0.0),
            'net_value': _safe_float(tx.get('net_value', 0), 0.0),
            'tx_time': tx_time,
        }
        if cycle_info:
            marker_item.update(cycle_info)
        trade_markers.append(marker_item)

    chart_data['trade_markers'] = trade_markers

    # 计算每个业绩曲线日期点的持有收益率（含已实现收益）
    net_values = chart_data.get('net_values', []) or []
    parsed_labels = resolved_label_dates

    tx_points = []
    for tx in transactions:
        tx_time = str(tx.get('tx_time', '')).strip()
        if not tx_time:
            continue
        tx_date_text = tx_time.split(' ')[0]
        try:
            tx_date = datetime.date.fromisoformat(tx_date_text)
        except Exception:
            continue
        tx_points.append({
            'date': tx_date,
            'tx_type': str(tx.get('tx_type', '')).strip().lower(),
            'shares': _safe_float(tx.get('shares', 0), 0.0),
            'amount': _safe_float(tx.get('amount', 0), 0.0),
        })

    tx_points.sort(key=lambda item: item['date'])
    tx_cursor = 0
    holding_shares = 0.0
    cumulative_buy = 0.0
    cumulative_sell = 0.0
    holding_return_pct = []

    for idx, label_date in enumerate(parsed_labels):
        if label_date is None:
            holding_return_pct.append(None)
            continue

        while tx_cursor < len(tx_points) and tx_points[tx_cursor]['date'] <= label_date:
            current_tx = tx_points[tx_cursor]
            tx_type = current_tx['tx_type']
            tx_shares = current_tx['shares']
            tx_amount = current_tx['amount']
            if tx_type == 'buy':
                holding_shares += tx_shares
                cumulative_buy += tx_amount
            elif tx_type == 'sell':
                holding_shares -= tx_shares
                cumulative_sell += tx_amount
            elif tx_type == 'dividend':
                cumulative_sell += tx_amount
            tx_cursor += 1

        if holding_shares < 0:
            holding_shares = 0.0

        if cumulative_buy <= 0:
            holding_return_pct.append(None)
            continue

        nav = _safe_float(net_values[idx] if idx < len(net_values) else None, None)
        if nav is None:
            holding_return_pct.append(None)
            continue

        total_value = cumulative_sell + holding_shares * nav
        total_return = total_value - cumulative_buy
        holding_return_pct.append(round(total_return / cumulative_buy * 100.0, 2))

    chart_data['holding_return_pct'] = holding_return_pct

    return jsonify({
        'chart_data': chart_data,
        'fund_info': {
            'code': fund_code,
            'name': fund_data['fund_name']
        }
    })


@app.route('/api/fund/profit-chart-data')
@login_required
def api_fund_profit_chart_data():
    """获取基金累计收益曲线数据（重写版：交易与净值按日对齐，提供更完整指标序列）。"""
    fund_code = request.args.get('code')
    date_interval = request.args.get('interval', DEFAULT_PROFIT_CHART_INTERVAL).strip().upper()
    if not fund_code:
        return jsonify({'error': 'Missing fund code'}), 400

    if date_interval not in PERFORMANCE_CHART_INTERVALS:
        return jsonify({'error': 'Invalid interval'}), 400

    user_id = get_current_user_id()
    user_funds = db.get_user_funds(user_id)
    if fund_code not in user_funds:
        return jsonify({'error': 'Fund not in user list'}), 400

    fund_data = {
        'fund_key': user_funds[fund_code]['fund_key'],
        'fund_name': user_funds[fund_code]['fund_name']
    }

    # 先构建“本地净值驱动”的业绩曲线数据：收益曲线与业绩曲线共享同一时间轴与净值基准。
    perf_data = _build_local_performance_chart_data(
        user_id=user_id,
        fund_code=fund_code,
        fund_name=fund_data['fund_name'],
        date_interval=date_interval,
        user_funds=user_funds,
    )

    labels = perf_data.get('labels', []) or []
    net_values = perf_data.get('net_values', []) or []
    growth_values = perf_data.get('growth', []) or []

    if len(net_values) < len(labels):
        net_values = net_values + [None] * (len(labels) - len(net_values))

    # 收益曲线同样以净值为基础：若当前区间净值覆盖不完整，则先触发净值补齐再重建曲线。
    interval_start_date = None
    establishment_date = _parse_iso_date(perf_data.get('establishment_date'))
    if date_interval == 'SINCE_ESTABLISHMENT' and isinstance(establishment_date, datetime.date):
        interval_start_date = establishment_date
    elif labels:
        interval_start_date = _parse_iso_date(labels[0])

    if isinstance(interval_start_date, datetime.date):
        interval_end_date = _nav_backfill_effective_end_date()
        interval_nav_map = db.get_fund_nav_history_range(
            fund_code,
            interval_start_date.isoformat(),
            interval_end_date.isoformat(),
        ) or {}

        interval_nav_dates = []
        for day in interval_nav_map.keys():
            day_obj = _parse_iso_date(day)
            if isinstance(day_obj, datetime.date):
                interval_nav_dates.append(day_obj)

        expected_interval_days = iter_cn_sse_trading_days(interval_start_date, interval_end_date)
        existing_interval_days = {
            day_obj for day_obj in interval_nav_dates
            if is_cn_sse_trading_day(day_obj)
        }
        missing_interval_days = [day for day in expected_interval_days if day not in existing_interval_days]
        request_segments = _build_backfill_request_segments(missing_interval_days, interval_end_date)

        if request_segments:
            wrote_count = 0
            for seg_start, seg_end in request_segments:
                if seg_start > seg_end:
                    continue
                remote_nav_map = _fetch_history_nav_map_by_date_range(
                    user_id,
                    fund_code,
                    seg_start.isoformat(),
                    seg_end.isoformat(),
                )
                if not remote_nav_map:
                    continue
                for nav_date, nav_value in remote_nav_map.items():
                    if db.upsert_fund_nav_history(fund_code, nav_date, nav_value, source='history_api_profit_interval_sync'):
                        wrote_count += 1

            if wrote_count > 0:
                logger.info(
                    f"收益曲线净值补齐完成【{fund_code}】: start={interval_start_date}, end={interval_end_date}, wrote={wrote_count}"
                )
                perf_data = _build_local_performance_chart_data(
                    user_id=user_id,
                    fund_code=fund_code,
                    fund_name=fund_data['fund_name'],
                    date_interval=date_interval,
                    user_funds=user_funds,
                )
                labels = perf_data.get('labels', []) or []
                net_values = perf_data.get('net_values', []) or []
                growth_values = perf_data.get('growth', []) or []
                if len(net_values) < len(labels):
                    net_values = net_values + [None] * (len(labels) - len(net_values))

    def _safe_float(value):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    transactions = db.get_fund_transactions(user_id, fund_code)
    sorted_txs = []
    for tx in transactions:
        tx_time = str(tx.get('tx_time', '') or '').strip()
        if not tx_time:
            continue

        tx_dt = None
        try:
            tx_dt = datetime.datetime.fromisoformat(tx_time.replace(' ', 'T'))
        except Exception:
            try:
                tx_dt = datetime.datetime.strptime(tx_time, '%Y-%m-%d %H:%M:%S')
            except Exception:
                pass
        if tx_dt is None:
            continue

        tx_type = str(tx.get('tx_type', '') or '').strip().lower()
        if tx_type not in ('buy', 'sell', 'dividend'):
            continue

        tx_amount = _safe_float(tx.get('amount'))
        tx_shares = _safe_float(tx.get('shares'))
        tx_fee = _safe_float(tx.get('fee'))
        tx_nav = _safe_float(tx.get('net_value'))

        sorted_txs.append({
            'datetime': tx_dt,
            'date': tx_dt.date(),
            'type': tx_type,
            'amount': max(tx_amount or 0.0, 0.0),
            'shares': max(tx_shares or 0.0, 0.0),
            'fee': max(tx_fee or 0.0, 0.0),
            'net_value': tx_nav if tx_nav and tx_nav > 0 else None,
        })

    sorted_txs.sort(key=lambda item: item['datetime'])

    resolved_label_dates = _resolve_curve_label_dates(labels)
    expected_dates = [item for item in resolved_label_dates if isinstance(item, datetime.date)]
    local_nav_map = db.get_fund_nav_history_range(fund_code) or {}

    has_valid_nav = any(_safe_float(value) not in (None, 0.0) for value in net_values)
    if labels:
        anchor_nav = None
        anchor_date = None

        # 优先使用当前区间内“最后一个真实净值点”作为锚点
        for idx in range(len(labels) - 1, -1, -1):
            nav_num = _safe_float(net_values[idx]) if idx < len(net_values) else None
            if nav_num is not None and nav_num > 0:
                anchor_nav = nav_num
                anchor_date = labels[idx]
                break

        # 其次使用交易净值锚点
        if anchor_nav is None or anchor_nav <= 0:
            for tx_item in reversed(sorted_txs):
                if tx_item['net_value'] is not None:
                    anchor_nav = tx_item['net_value']
                    anchor_date = tx_item['date'].isoformat()
                    break

        # 再其次使用本地净值历史锚点
        if (anchor_nav is None or anchor_nav <= 0) and local_nav_map:
            last_date = sorted(local_nav_map.keys())[-1]
            local_anchor = _safe_float(local_nav_map.get(last_date))
            if local_anchor is not None and local_anchor > 0:
                anchor_nav = local_anchor
                anchor_date = last_date

        # 最后回退到最新行情净值锚点
        if anchor_nav is None or anchor_nav <= 0:
            latest_nav, latest_nav_date, _ = _get_latest_fund_quote(user_id, fund_code)
            latest_nav_num = _safe_float(latest_nav)
            if latest_nav_num is not None and latest_nav_num > 0:
                anchor_nav = latest_nav_num
                anchor_date = latest_nav_date

        reference_index = None
        if anchor_date:
            try:
                anchor_date_obj = datetime.date.fromisoformat(str(anchor_date))
            except Exception:
                anchor_date_obj = None
            if anchor_date_obj is not None:
                for idx, point_date in enumerate(resolved_label_dates):
                    if point_date == anchor_date_obj:
                        reference_index = idx
                        break

            if reference_index is None:
                anchor_date_text = str(anchor_date)
                anchor_suffix = anchor_date_text[5:] if len(anchor_date_text) >= 10 else anchor_date_text
                for idx, label in enumerate(labels):
                    label_text = str(label)
                    if label_text == anchor_date_text or label_text.endswith(anchor_suffix):
                        reference_index = idx
                        break

        if reference_index is None:
            for idx in range(len(growth_values) - 1, -1, -1):
                if _safe_float(growth_values[idx]) is not None:
                    reference_index = idx
                    break

        # 使用“锚点净值 + 同区间涨幅”补齐缺失净值点（非插值，不跨日平铺）。
        # 真实净值始终优先，重建值仅用于缺失补齐，避免污染已落库的真实历史净值。
        if anchor_nav and anchor_nav > 0 and reference_index is not None:
            reference_growth = _safe_float(growth_values[reference_index]) or 0.0
            denominator = 1 + reference_growth / 100.0
            if abs(denominator) > 1e-8:
                base_nav = anchor_nav / denominator
                rebuilt_nav_values = []
                for idx, growth in enumerate(growth_values):
                    growth_num = _safe_float(growth)
                    if growth_num is None:
                        rebuilt_nav_values.append(None)
                    else:
                        rebuilt_nav_values.append(round(base_nav * (1 + growth_num / 100.0), 4))

                # 真实净值优先；仅在缺失时用重建值补齐
                max_len = max(len(net_values), len(rebuilt_nav_values))
                merged_nav_values = []
                for idx in range(max_len):
                    real_nav = _safe_float(net_values[idx]) if idx < len(net_values) else None
                    rebuilt_nav = _safe_float(rebuilt_nav_values[idx]) if idx < len(rebuilt_nav_values) else None
                    if real_nav is not None and real_nav > 0:
                        merged_nav_values.append(round(real_nav, 4))
                    elif rebuilt_nav is not None and rebuilt_nav > 0:
                        merged_nav_values.append(round(rebuilt_nav, 4))
                    else:
                        merged_nav_values.append(None)
                net_values = merged_nav_values

    nav_by_date = {}
    for idx, point_date in enumerate(resolved_label_dates):
        if point_date is None:
            continue
        nav_value = _safe_float(net_values[idx]) if idx < len(net_values) else None
        if nav_value is not None and nav_value > 0:
            nav_by_date[point_date] = round(nav_value, 4)

    if not nav_by_date and sorted_txs:
        for tx in sorted_txs:
            if tx['net_value'] is not None:
                nav_by_date[tx['date']] = round(tx['net_value'], 4)

    expanded_dates = sorted(nav_by_date.keys()) if nav_by_date else sorted(expected_dates)
    expanded_labels = [item.isoformat() for item in expanded_dates]
    expanded_net_values = [nav_by_date.get(item) for item in expanded_dates]

    tx_index = 0
    cumulative_buy = 0.0
    cumulative_sell = 0.0
    cumulative_dividend = 0.0
    realized_gain = 0.0
    holding_shares = 0.0
    remaining_cost = 0.0

    profit_values = []
    holding_gain_values = []
    realized_gain_values = []
    position_value_values = []
    remaining_cost_values = []
    cumulative_buy_values = []
    cumulative_sell_values = []
    cumulative_dividend_values = []
    profit_rate_values = []

    for idx, point_date in enumerate(expanded_dates):
        
        while tx_index < len(sorted_txs) and point_date and sorted_txs[tx_index]['date'] <= point_date:
            tx = sorted_txs[tx_index]
            tx_type = tx['type']
            tx_amount = tx['amount']
            tx_shares = tx['shares']
            tx_fee = tx['fee']

            if tx_type == 'buy' and tx_shares > 0:
                total_cost = tx_amount + tx_fee
                cumulative_buy += total_cost
                holding_shares += tx_shares
                remaining_cost += total_cost
            elif tx_type == 'sell' and tx_shares > 0 and holding_shares > 1e-8:
                sold_shares = min(tx_shares, holding_shares)
                avg_cost = (remaining_cost / holding_shares) if holding_shares > 1e-8 else 0.0
                sold_cost = sold_shares * avg_cost
                proceeds = max(tx_amount - tx_fee, 0.0)

                if tx_shares > sold_shares and tx_shares > 0:
                    proceeds *= (sold_shares / tx_shares)

                cumulative_sell += proceeds
                realized_gain += (proceeds - sold_cost)
                remaining_cost = max(remaining_cost - sold_cost, 0.0)
                holding_shares = max(holding_shares - sold_shares, 0.0)
            elif tx_type == 'dividend':
                cumulative_dividend += tx_amount
                cumulative_sell += tx_amount
                realized_gain += tx_amount

            tx_index += 1

        net_value = expanded_net_values[idx] if idx < len(expanded_net_values) else None
        net_value_num = _safe_float(net_value)

        if net_value_num is None and holding_shares > 1e-8:
            position_value = None
            holding_gain = None
            total_profit = None
        else:
            position_value = (holding_shares * (net_value_num or 0.0))
            holding_gain = position_value - remaining_cost
            total_profit = realized_gain + holding_gain

        invested_base = cumulative_buy if cumulative_buy > 1e-8 else 0.0
        if total_profit is None or invested_base <= 0:
            profit_rate = None
        else:
            profit_rate = total_profit / invested_base * 100.0

        profit_values.append(round(total_profit, 2) if total_profit is not None else None)
        holding_gain_values.append(round(holding_gain, 2) if holding_gain is not None else None)
        realized_gain_values.append(round(realized_gain, 2))
        position_value_values.append(round(position_value, 2) if position_value is not None else None)
        remaining_cost_values.append(round(remaining_cost, 2))
        cumulative_buy_values.append(round(cumulative_buy, 2))
        cumulative_sell_values.append(round(cumulative_sell, 2))
        cumulative_dividend_values.append(round(cumulative_dividend, 2))
        profit_rate_values.append(round(profit_rate, 4) if profit_rate is not None else None)

    latest_profit = next((value for value in reversed(profit_values) if value is not None), None)
    latest_profit_rate = next((value for value in reversed(profit_rate_values) if value is not None), None)

    chart_data = {
        'labels': expanded_labels,
        'profit_values': profit_values,
        'holding_gain_values': holding_gain_values,
        'realized_gain_values': realized_gain_values,
        'position_value_values': position_value_values,
        'remaining_cost_values': remaining_cost_values,
        'cumulative_buy_values': cumulative_buy_values,
        'cumulative_sell_values': cumulative_sell_values,
        'cumulative_dividend_values': cumulative_dividend_values,
        'profit_rate_values': profit_rate_values,
        'date_interval': perf_data.get('date_interval', date_interval),
        'interval_label': perf_data.get('interval_label', date_interval),
        'available_intervals': perf_data.get('available_intervals', []),
        'establishment_date': perf_data.get('establishment_date'),
        'latest_profit': latest_profit,
        'latest_profit_rate': latest_profit_rate,
    }

    return jsonify({
        'chart_data': chart_data,
        'fund_info': {
            'code': fund_code,
            'name': fund_data['fund_name']
        }
    })


@app.route('/api/fund/chart-default', methods=['POST'])
@login_required
def api_fund_chart_default():
    """设置估值趋势图默认基金"""
    data = request.json
    fund_code = data.get('fund_code')
    if not fund_code:
        return jsonify({'error': 'Missing fund code'}), 400

    user_id = get_current_user_id()
    user_funds = db.get_user_funds(user_id)

    if fund_code not in user_funds:
        return jsonify({'error': 'Fund not in user list'}), 400

    db.update_chart_default(user_id, fund_code)
    return jsonify({'success': True})


@app.route('/api/config/refresh', methods=['GET'])
def api_config_refresh():
    """获取页面刷新配置"""
    config = get_page_refresh_config()
    return jsonify(config)


@app.route('/sectors', methods=['GET'])
@login_required
def get_sectors():
    """行业板块基金查询页面"""
    user_id = get_current_user_id()
    importlib.reload(fund)
    my_fund = fund.LanFund(user_id=user_id, db=db)

    # 加载行业板块数据
    try:
        sectors_content = my_fund.bk_html()
        logger.debug("✓ 行业板块")
    except Exception as e:
        sectors_content = f"<p style='color:#f44336;'>数据加载失败: {str(e)}</p>"

    # 加载板块基金查询数据
    try:
        select_fund_content = my_fund.select_fund_html()
        logger.debug("✓ 板块基金查询")
    except Exception as e:
        select_fund_content = f"<p style='color:#f44336;'>数据加载失败: {str(e)}</p>"

    from src.module_html import get_sectors_page_html
    html = get_sectors_page_html(
        sectors_content=sectors_content,
        select_fund_content=select_fund_content,
        fund_map=my_fund.CACHE_MAP,
        username=get_current_username()
    )
    return html


class FilteredWSGIRequestLogger:
    """
    WSGI中间件，过滤静态资源请求日志（如/static/、/favicon.ico等）。
    """
    def __init__(self, app):
        self.app = app

    def __call__(self, environ, start_response):
        path = environ.get('PATH_INFO', '')
        # 过滤静态资源和favicon请求
        if path.startswith('/static/') or path.endswith('.ico') or path.startswith('/favicon'):
            # 禁止Flask默认日志输出
            import logging
            logging.getLogger('werkzeug').setLevel(logging.ERROR)
        else:
            import logging
            logging.getLogger('werkzeug').setLevel(logging.INFO)
        return self.app(environ, start_response)

if __name__ == '__main__':
    # 用中间件包裹Flask app，过滤静态资源日志
    from werkzeug.serving import run_simple
    app.wsgi_app = FilteredWSGIRequestLogger(app.wsgi_app)

    # 仅在 reloader 子进程启动后台日志清理线程，避免父进程重复启动
    if os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        _start_log_cleanup_worker_if_needed()

    # 避免日志文件/本地数据库写入触发 reloader 反复重启（会造成“日志清理频繁触发”的错觉）。
    # 低版本 Werkzeug 可能不支持 exclude_patterns，故做兼容回退。
    try:
        run_simple(
            '0.0.0.0',
            8311,
            app,
            use_reloader=True,
            use_debugger=False,
            exclude_patterns=[
                'cache/logs/*',
                'cache/*.db',
                'cache/**/*.db',
                '*.db-journal',
                '*.db-wal',
                '*.db-shm',
            ],
        )
    except TypeError:
        run_simple('0.0.0.0', 8311, app, use_reloader=True, use_debugger=False)
