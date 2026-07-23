# -*- coding: UTF-8 -*-
"""Import service — Excel transaction import with async job tracking."""

import datetime
import io
import json
import os
import threading
import time
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from loguru import logger

from src.services.metrics import safe_float, quantize_shares_2, calculate_holding_shares_by_time

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

IMPORT_JOB_STORE = {}
IMPORT_JOB_LOCK = threading.Lock()
IMPORT_DETAIL_LOG_PATH = str(Path(os.environ.get("FUNDEVAL_LOG_DIR", _PROJECT_ROOT / "cache" / "logs")) / "transaction_import.log")


# ── job state helpers ──────────────────────────────────────────────────

def _set_import_job_state(job_id, **kwargs):
    with IMPORT_JOB_LOCK:
        state = IMPORT_JOB_STORE.get(job_id, {}).copy()
        state.update(kwargs)
        state['updated_at'] = time.time()
        IMPORT_JOB_STORE[job_id] = state
        return state


def _get_import_job_state(job_id):
    with IMPORT_JOB_LOCK:
        state = IMPORT_JOB_STORE.get(job_id)
        return dict(state) if state else None


def _cleanup_finished_import_jobs(max_age_seconds=1800):
    now_ts = time.time()
    with IMPORT_JOB_LOCK:
        expired_ids = [
            job_id for job_id, state in IMPORT_JOB_STORE.items()
            if state.get('done') and (now_ts - float(state.get('updated_at', now_ts))) > max_age_seconds
        ]
        for job_id in expired_ids:
            IMPORT_JOB_STORE.pop(job_id, None)


def _append_import_detail_log(level, message, **fields):
    try:
        ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        detail_parts = []
        for key, value in fields.items():
            if value is None or value == '':
                continue
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            detail_parts.append(f'{key}={value}')
        detail_text = ' | '.join(detail_parts)
        line = f'[{ts}] [{level}] {message}'
        if detail_text:
            line += f' | {detail_text}'
        with open(IMPORT_DETAIL_LOG_PATH, 'a', encoding='utf-8') as f:
            f.write(line + '\n')
    except Exception as e:
        logger.error(f"写入交易导入本地日志失败: {e}")


# ── import parsing helpers ─────────────────────────────────────────────

def _normalize_import_tx_type(tx_type_raw):
    tx_type_text = str(tx_type_raw or '').strip().lower()
    if not tx_type_text:
        return None

    normalized_text = tx_type_text.replace(' ', '')
    buy_keywords = ('buy', '买入', '申购', '购入')
    sell_keywords = ('sell', '卖出', '赎回', '卖')
    dividend_keywords = ('dividend', '分红', '现金分红', '派息', '红利')

    if any(keyword in normalized_text for keyword in buy_keywords):
        return 'buy'
    if any(keyword in normalized_text for keyword in sell_keywords):
        return 'sell'
    if any(keyword in normalized_text for keyword in dividend_keywords):
        return 'dividend'
    return None


def _parse_import_trade_datetime(raw_value):
    if isinstance(raw_value, datetime.datetime):
        return raw_value.replace(microsecond=0)

    if isinstance(raw_value, datetime.date):
        return datetime.datetime.combine(raw_value, datetime.time(0, 0, 0))

    if isinstance(raw_value, (int, float)):
        try:
            excel_base = datetime.datetime(1899, 12, 30)
            days = float(raw_value)
            seconds = int(round((days - int(days)) * 24 * 3600))
            return (excel_base + datetime.timedelta(days=int(days), seconds=seconds)).replace(microsecond=0)
        except Exception:
            return None

    text = str(raw_value or '').strip()
    if not text:
        return None

    parse_formats = [
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%Y/%m/%d %H:%M:%S',
        '%Y/%m/%d %H:%M',
        '%Y-%m-%d',
        '%Y/%m/%d',
    ]
    for fmt in parse_formats:
        try:
            parsed = datetime.datetime.strptime(text, fmt)
            return parsed.replace(microsecond=0)
        except Exception:
            continue

    try:
        return datetime.datetime.fromisoformat(text.replace('T', ' ')).replace(microsecond=0)
    except Exception:
        return None


def _normalize_import_text(value):
    if value is None:
        return ''
    text = str(value)
    text = text.replace('\r', '').replace('\n', '').strip()
    return text


def _read_excel_rows_for_transaction_import(file_storage):
    if load_workbook is None:
        return None, '缺少 openpyxl 依赖，无法解析Excel，请安装 openpyxl'

    try:
        workbook = load_workbook(file_storage, data_only=True, read_only=True)
    except Exception as e:
        return None, f'Excel解析失败: {e}'

    sheet = workbook.active
    if sheet is None:
        return None, 'Excel解析失败: 未找到可用工作表'
    rows = list(sheet.iter_rows(values_only=False))
    if not rows:
        return None, 'Excel内容为空'

    header_row = rows[0]
    normalized_headers = [str(getattr(cell, 'value', '') or '').strip().lower() for cell in header_row]
    alias_map = {
        'order_no': {'order_no', '订单号', '委托编号', '订单编号', '交易单号'},
        'fund_code': {'fund_code', '基金代码', '代码', '基金'},
        'tx_type': {'tx_type', '交易类型', '类型', '买卖方向', '买卖'},
        'trade_time': {'trade_time', '交易时间', '成交时间', '确认时间', '日期'},
        'amount': {'amount', '确认金额', '金额', '到账金额', '成交金额'},
        'confirmed_shares': {'confirmed_shares', '确认份额', '份额', '成交份额', '确认数量'},
        'fee': {'fee', '手续费', '费用', '交易手续费'},
    }

    col_index = {}
    for field_name, aliases in alias_map.items():
        for idx, header in enumerate(normalized_headers):
            if header in {str(alias).strip().lower() for alias in aliases}:
                col_index[field_name] = idx
                break

    required_fields = ['order_no', 'fund_code', 'tx_type', 'trade_time', 'amount']
    for field_name in required_fields:
        if field_name not in col_index:
            return None, f'Excel缺少必填列: {field_name}'

    parsed_rows = []
    for row_num, row_cells in enumerate(rows[1:], start=2):
        row_dict = {'row_num': row_num}
        for field_name, idx in col_index.items():
            cell = row_cells[idx] if idx < len(row_cells) else None
            raw_value = getattr(cell, 'value', None) if cell is not None else None

            if field_name == 'fund_code':
                if raw_value is None:
                    row_dict[field_name] = None
                elif isinstance(raw_value, str):
                    row_dict[field_name] = _normalize_import_text(raw_value)
                elif isinstance(raw_value, (int, float)):
                    if isinstance(raw_value, float):
                        if raw_value.is_integer():
                            code_text = str(int(raw_value))
                        else:
                            code_text = _normalize_import_text(raw_value)
                    else:
                        code_text = str(int(raw_value))

                    number_format = str(getattr(cell, 'number_format', '') or '').strip()
                    if number_format and set(number_format) == {'0'}:
                        code_text = code_text.zfill(len(number_format))
                    elif code_text.isdigit() and len(code_text) < 6:
                        code_text = code_text.zfill(6)

                    row_dict[field_name] = code_text
                else:
                    row_dict[field_name] = _normalize_import_text(raw_value)
            else:
                if isinstance(raw_value, str):
                    row_dict[field_name] = _normalize_import_text(raw_value)
                else:
                    row_dict[field_name] = raw_value
        parsed_rows.append(row_dict)

    return parsed_rows, None


def _log_import_share_mismatch(fund_code, trade_dt, amount, tx_type, net_value, confirmed_shares, computed_shares):
    tx_label_map = {
        'buy': '买入',
        'sell': '卖出',
    }
    trade_text = trade_dt.strftime('%Y-%m-%d') if isinstance(trade_dt, datetime.datetime) else str(trade_dt or '-')
    tx_label = tx_label_map.get(str(tx_type or '').lower(), str(tx_type or '-'))
    message = (
        f'{fund_code}基金于{trade_text}日的{float(amount or 0):.2f}元{tx_label}交易，'
        f'净值{float(net_value or 0):.4f}，确认份额{float(confirmed_shares or 0):.2f}与计算份额{float(computed_shares or 0):.2f}不一致'
    )
    logger.warning('交易导入份额校验不一致: fund_code={}, tx_type={}', fund_code, tx_type)
    _append_import_detail_log(
        'WARNING',
        '交易导入份额校验不一致',
        fund_code=fund_code,
        tx_type=tx_label,
    )
    return message


# ── ImportService ──────────────────────────────────────────────────────

class ImportService:
    """Service for asynchronous Excel transaction import and progress polling."""

    def __init__(self, fund_repo, transaction_repo, nav_repo, get_lan_fund_func, transaction_service):
        self._fund_repo = fund_repo
        self._transaction_repo = transaction_repo
        self._nav_repo = nav_repo
        self._get_lan_fund = get_lan_fund_func
        self._transaction_service = transaction_service

    def import_transactions(self, user_id, file_bytes, filename):
        """Start an async import job. Returns job metadata dict."""
        _cleanup_finished_import_jobs()

        lower_name = str(filename or '').lower()
        if not lower_name.endswith('.xlsx'):
            return {'success': False, 'message': '仅支持 .xlsx 格式Excel文件'}

        if not file_bytes:
            return {'success': False, 'message': '上传文件为空'}

        job_id = _generate_job_id()
        _set_import_job_state(
            job_id,
            user_id=user_id,
            success=None,
            done=False,
            status='queued',
            title='交易记录导入中',
            detail='文件上传完成，等待服务器处理...',
            percent=0,
            processed_count=0,
            total_count=0,
            imported_count=0,
            failed_count=0,
            duplicate_count=0,
            warning_messages=[],
            failed_rows=[],
            message='导入任务已创建',
        )

        worker = threading.Thread(
            target=self._run_transaction_import_job,
            args=(job_id, user_id, file_bytes),
            daemon=True,
        )
        worker.start()

        return {
            'success': True,
            'accepted': True,
            'job_id': job_id,
            'message': '文件上传成功，已开始导入',
        }

    def get_import_progress(self, user_id, job_id):
        """Poll import job progress."""
        state = _get_import_job_state(job_id)
        if not state or int(state.get('user_id', -1)) != int(user_id):
            return {'success': False, 'message': '导入任务不存在或已过期'}
        return {'success': True, 'job': state}

    def _run_transaction_import_job(self, job_id, user_id, file_bytes):
        try:
            _set_import_job_state(job_id, status='parsing', detail='正在解析Excel文件...', percent=5)
            parsed_rows, read_error = _read_excel_rows_for_transaction_import(io.BytesIO(file_bytes))
            if read_error:
                _append_import_detail_log('ERROR', f'Excel解析失败: {read_error}', job_id=job_id)
                _set_import_job_state(
                    job_id,
                    success=False,
                    done=True,
                    status='failed',
                    detail='Excel解析失败',
                    percent=100,
                    message=read_error,
                    failed_count=1,
                )
                return

            if not parsed_rows:
                _append_import_detail_log('ERROR', 'Excel中无可导入数据', job_id=job_id)
                _set_import_job_state(
                    job_id,
                    success=False,
                    done=True,
                    status='failed',
                    detail='Excel中无可导入数据',
                    percent=100,
                    message='Excel中无可导入数据',
                    failed_count=1,
                )
                return

            my_fund = self._get_lan_fund(user_id=user_id)
            user_funds = self._fund_repo.get_user_funds(user_id)

            imported_count = 0
            duplicate_count = 0
            failed_rows = []
            warning_messages = []

            preprocessed_rows = []
            for row in parsed_rows:
                parsed_trade_dt = _parse_import_trade_datetime(row.get('trade_time'))
                preprocessed = dict(row)
                preprocessed['parsed_trade_dt'] = parsed_trade_dt
                preprocessed_rows.append(preprocessed)

            sorted_rows = sorted(
                preprocessed_rows,
                key=lambda row: (
                    row.get('parsed_trade_dt') or datetime.datetime.max,
                    int(row.get('row_num', 0) or 0),
                )
            )
            total_count = len(sorted_rows)

            nav_cache = {}
            order_exists_cache = {}
            seen_order_nos_in_file = set()

            def _get_cached_trade_nav(fund_code, trade_dt):
                if not isinstance(trade_dt, datetime.datetime):
                    return None, None
                start_date = trade_dt.date() if trade_dt.time() < datetime.time(15, 0, 0) else (trade_dt.date() + datetime.timedelta(days=1))
                cache_key = (fund_code, start_date.isoformat())
                if cache_key in nav_cache:
                    return nav_cache[cache_key]
                resolved = self._transaction_service._resolve_net_value_for_trade_datetime(user_id, fund_code, trade_dt)
                nav_cache[cache_key] = resolved
                return resolved

            _set_import_job_state(
                job_id,
                status='processing',
                detail=f'服务器处理中：已处理 0/{total_count}',
                total_count=total_count,
                percent=10,
            )

            for index, item in enumerate(sorted_rows, start=1):
                row_num = int(item.get('row_num', 0) or 0)
                row_diag = {}
                try:
                    order_no = _normalize_import_text(item.get('order_no', ''))
                    code = _normalize_import_text(item.get('fund_code', ''))
                    tx_type = _normalize_import_tx_type(item.get('tx_type'))
                    fee = safe_float(item.get('fee', 0), 0.0)
                    trade_dt = item.get('parsed_trade_dt')

                    amount_raw = item.get('amount', None)
                    amount = safe_float(amount_raw, None)
                    confirmed_shares_raw = item.get('confirmed_shares', None)
                    confirmed_shares = safe_float(confirmed_shares_raw, None)

                    if not order_no:
                        raise ValueError('订单号不能为空')

                    if order_no in seen_order_nos_in_file:
                        duplicate_count += 1
                        continue

                    exists_in_db = order_exists_cache.get(order_no)
                    if exists_in_db is None:
                        exists_in_db = self._transaction_repo.exists_transaction_order_no(user_id, order_no)
                        order_exists_cache[order_no] = exists_in_db
                    if exists_in_db:
                        duplicate_count += 1
                        continue

                    seen_order_nos_in_file.add(order_no)

                    if not code:
                        raise ValueError('基金代码不能为空')
                    if tx_type not in ('buy', 'sell', 'dividend'):
                        raise ValueError('交易类型仅支持买入/卖出/分红')
                    if trade_dt is None:
                        raise ValueError('交易时间格式错误')
                    if amount is None or amount <= 0:
                        raise ValueError('确认金额必须大于0')
                    if fee is None or fee < 0:
                        raise ValueError('手续费必须大于等于0')

                    if code not in user_funds:
                        my_fund.add_code(code)
                        user_funds = self._fund_repo.get_user_funds(user_id)
                    if code not in user_funds:
                        raise ValueError(f'基金代码无效或添加失败: {code}')

                    db_snapshot_shares = float(user_funds.get(code, {}).get('shares', 0) or 0)
                    row_diag['db_snapshot_shares'] = round(db_snapshot_shares, 6)

                    net_value = None
                    if tx_type in ('buy', 'sell'):
                        net_value, _nav_date = _get_cached_trade_nav(code, trade_dt)
                        if net_value is None or net_value <= 0:
                            raise ValueError('未查询到可用净值')
                        row_diag['net_value'] = round(float(net_value), 6)

                    confirm_amount = float(Decimal(str(amount)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    fee = float(Decimal(str(fee)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                    row_diag['confirm_amount'] = round(confirm_amount, 2)
                    row_diag['fee'] = round(fee, 2)

                    if tx_type == 'buy':
                        buy_base_amount = float((Decimal(str(confirm_amount)) - Decimal(str(fee))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                        if buy_base_amount <= 0:
                            raise ValueError('买入确认金额必须大于手续费')
                        row_diag['buy_base_amount'] = round(buy_base_amount, 2)
                        computed_shares = float((Decimal(str(buy_base_amount)) / Decimal(str(net_value))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                        if computed_shares <= 0:
                            raise ValueError('买入确认份额计算结果为0')
                        row_diag['computed_shares'] = round(computed_shares, 6)
                        shares = computed_shares
                        if confirmed_shares is not None and confirmed_shares > 0:
                            confirmed_shares = quantize_shares_2(confirmed_shares)
                            row_diag['confirmed_shares'] = confirmed_shares
                            if abs(confirmed_shares - computed_shares) > 1e-8:
                                warning_messages.append(
                                    _log_import_share_mismatch(code, trade_dt, confirm_amount, tx_type, net_value, confirmed_shares, computed_shares)
                                )
                            shares = confirmed_shares
                        new_shares = self._fund_repo.update_fund_shares_delta(user_id, code, shares)
                        if new_shares is None:
                            db_actual_shares = float(self._fund_repo.get_user_funds(user_id).get(code, {}).get('shares', 0) or 0)
                            raise ValueError(
                                f'买入份额更新失败（数据库持仓{db_actual_shares:.2f}, 本次买入{shares:.2f}, '
                                f'确认金额{confirm_amount:.2f}, 手续费{fee:.2f}, 净值{net_value:.4f}）'
                            )
                    elif tx_type == 'sell':
                        sell_gross_amount = float((Decimal(str(confirm_amount)) + Decimal(str(fee))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                        if sell_gross_amount <= 0:
                            raise ValueError('卖出金额与手续费之和必须大于0')
                        row_diag['sell_gross_amount'] = round(sell_gross_amount, 2)
                        computed_shares = float((Decimal(str(sell_gross_amount)) / Decimal(str(net_value))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
                        if computed_shares <= 0:
                            raise ValueError('卖出确认份额计算结果为0')
                        row_diag['computed_shares'] = round(computed_shares, 6)
                        shares = computed_shares
                        if confirmed_shares is not None and confirmed_shares > 0:
                            confirmed_shares = quantize_shares_2(confirmed_shares)
                            row_diag['confirmed_shares'] = confirmed_shares
                            if abs(confirmed_shares - computed_shares) > 1e-8:
                                warning_messages.append(
                                    _log_import_share_mismatch(code, trade_dt, confirm_amount, tx_type, net_value, confirmed_shares, computed_shares)
                                )
                            shares = confirmed_shares

                        current_holding = calculate_holding_shares_by_time(self._transaction_repo, user_id, code, up_to_dt=trade_dt)
                        row_diag['timeline_holding'] = round(current_holding, 6)
                        request_shares_2 = quantize_shares_2(shares)
                        available_shares_2 = quantize_shares_2(current_holding)
                        row_diag['request_shares_2'] = request_shares_2
                        row_diag['available_shares_2'] = available_shares_2
                        if request_shares_2 > available_shares_2:
                            raise ValueError(
                                f'卖出份额超过交易时点可用持仓（可用{available_shares_2:.2f}, 本次{request_shares_2:.2f}, '
                                f'确认金额{confirm_amount:.2f}, 手续费{fee:.2f}, 净值{net_value:.4f}）'
                            )

                        shares = request_shares_2
                        row_diag['executed_shares_2'] = shares

                        recalculated = self._fund_repo.recalculate_fund_shares_from_transactions(user_id, code)
                        current_total_shares = float((recalculated or {}).get('current_shares', db_snapshot_shares) or 0)
                        row_diag['current_total_shares'] = round(current_total_shares, 6)
                        if shares > quantize_shares_2(current_total_shares):
                            raise ValueError(
                                f'卖出份额超过当前总持仓（当前{current_total_shares:.2f}, 本次{shares:.2f}）'
                            )
                        if code not in user_funds:
                            user_funds[code] = {}
                        user_funds[code]['shares'] = float(current_total_shares)

                        new_shares = self._fund_repo.update_fund_shares_delta(user_id, code, -shares)
                        if new_shares is None:
                            db_actual_shares = float(self._fund_repo.get_user_funds(user_id).get(code, {}).get('shares', 0) or 0)
                            raise ValueError(
                                f'卖出份额更新失败（数据库持仓{db_actual_shares:.2f}, 交易时点可用{current_holding:.2f}, '
                                f'本次卖出{shares:.2f}, 确认金额{confirm_amount:.2f}, 手续费{fee:.2f}, 净值{net_value:.4f}）'
                            )
                    else:
                        shares = 0.0
                        new_shares = float(user_funds.get(code, {}).get('shares', 0) or 0)
                        row_diag['dividend_amount'] = round(confirm_amount, 2)

                    tx_time = trade_dt.strftime('%Y-%m-%d %H:%M:%S')
                    tx_id = self._transaction_repo.add_fund_transaction(
                        user_id=user_id,
                        fund_code=code,
                        tx_type=tx_type,
                        amount=confirm_amount,
                        shares=shares,
                        net_value=net_value,
                        tx_time=tx_time,
                        fee=fee,
                        order_no=order_no,
                    )
                    if tx_id is None:
                        if tx_type == 'buy':
                            self._fund_repo.update_fund_shares_delta(user_id, code, -shares)
                        elif tx_type == 'sell':
                            self._fund_repo.update_fund_shares_delta(user_id, code, shares)
                        raise ValueError('交易记录写入失败')

                    imported_count += 1
                    if code not in user_funds:
                        user_funds[code] = {}
                    user_funds[code]['shares'] = float(new_shares or 0)
                    order_exists_cache[order_no] = True
                except Exception as row_error:
                    row_code = _normalize_import_text(item.get('fund_code', '')) or '-'
                    row_order_no = _normalize_import_text(item.get('order_no', '')) or '-'
                    row_tx_type = _normalize_import_text(item.get('tx_type', '')) or '-'
                    row_amount = _normalize_import_text(item.get('amount', '')) or '-'
                    row_fee = _normalize_import_text(item.get('fee', '')) or '-'
                    parsed_trade_dt = item.get('parsed_trade_dt')
                    if isinstance(parsed_trade_dt, datetime.datetime):
                        row_trade_date = parsed_trade_dt.strftime('%Y-%m-%d')
                    else:
                        row_trade_date = _normalize_import_text(item.get('trade_time', '')) or '-'
                    logger.warning(
                        '交易导入失败: row={}, fund_code={}, tx_type={}, error_type={}',
                        row_num, row_code, row_tx_type, type(row_error).__name__,
                    )
                    _append_import_detail_log(
                        'ERROR',
                        '交易导入失败',
                        row=row_num,
                        fund_code=row_code,
                        tx_type=row_tx_type,
                        error_type=type(row_error).__name__,
                    )
                    failed_rows.append({
                        'row': row_num,
                        'reason': f"基金{row_code} 日期{row_trade_date}：{str(row_error)}",
                    })
                finally:
                    if index % 20 == 0 or index == total_count:
                        process_percent = 10 + int((index / max(total_count, 1)) * 85)
                        _set_import_job_state(
                            job_id,
                            status='processing',
                            detail=f'服务器处理中：已处理 {index}/{total_count}',
                            percent=min(process_percent, 95),
                            processed_count=index,
                            total_count=total_count,
                            imported_count=imported_count,
                            failed_count=len(failed_rows),
                            duplicate_count=duplicate_count,
                            warning_messages=warning_messages[-20:],
                        )

            failed_count = len(failed_rows)
            success = imported_count > 0
            message = (
                f'导入完成：成功 {imported_count} 条，失败 {failed_count} 条，重复订单 {duplicate_count} 条'
                if success else
                f'导入失败，共 {failed_count} 条错误，重复订单 {duplicate_count} 条'
            )
            _set_import_job_state(
                job_id,
                success=success,
                done=True,
                status='completed' if success else 'failed',
                detail=f'服务器处理中：已处理 {total_count}/{total_count}',
                percent=100,
                processed_count=total_count,
                total_count=total_count,
                imported_count=imported_count,
                failed_count=failed_count,
                duplicate_count=duplicate_count,
                warning_messages=warning_messages[-20:],
                failed_rows=failed_rows,
                message=message,
            )
        except Exception as e:
            logger.error(f"交易记录导入后台任务失败: {e}")
            _append_import_detail_log('ERROR', f'交易记录导入后台任务失败: {e}', job_id=job_id)
            state = _get_import_job_state(job_id) or {}
            total_count = int(state.get('total_count', 0) or 0)
            processed_count = int(state.get('processed_count', 0) or 0)
            _set_import_job_state(
                job_id,
                success=False,
                done=True,
                status='failed',
                detail=f'服务器处理中：已处理 {processed_count}/{total_count}' if total_count else '服务器处理失败',
                percent=100,
                message=f'导入交易记录失败: {str(e)}',
            )


def _generate_job_id():
    import uuid
    return uuid.uuid4().hex
